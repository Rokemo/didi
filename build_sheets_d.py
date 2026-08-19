# -*- coding: utf-8 -*-
"""11合同冗余 —— 超期未交付占压分析 + 接单判断
逻辑口径对齐参考表 D:\\倍优待交付情况6.2.xlsx（汇总6.2 / 90天明细 / 总明细）
数据源统一为「原始数据表」，销售预测由「销售预测表」用户维护。
"""
from openpyxl.utils import get_column_letter as CL
from openpyxl.styles import Font
from openpyxl.formatting.rule import FormulaRule, DataBarRule
from build_part1 import *
from build_util import *

DATE_FMT = "yyyy-mm-dd"

Q_RAW = q(SH_RAW)
Q_FCST = q(SH_FCST)
Q_REDUN = q(SH_REDUN)

RD_NB = 600     # 辅助计算区行缓冲（原始数据表可增长到 600 行而无需改公式）
RD_R0 = 11      # 各汇总区 / 辅助区 数据起始行
RD_H0 = 48      # 辅助计算区起始列 = AV

# 固定 6 档（不随阈值参数变化，保证 SUMIF 文本精确匹配）
RD_BUCKETS = [
    ("未到期",      "正常",     "尚未到交货日，属在途正常单"),
    ("逾期1-90",    "关注",     "轻度逾期，常规催办即可"),
    ("逾期91-180",  "需关注",   "已构成冗余，须书面确认新交期"),
    ("逾期181-270", "跟踪",     "长期占压，评估改单/转产/让步接收"),
    ("逾期271-365", "跟踪",     "严重占压，启动供应商问责"),
    ("逾期365以上", "严重呆滞", "启动清尾：核销、退单或转内销"),
]


def build_redun(wb, dims):
    """dims: redun_dims() 结果（groups/facs/skus/combos，均按未到货金额降序）"""
    ws = wb.create_sheet(SH_REDUN)
    R1, RN = RD_R0, RD_R0 + RD_NB - 1                                  # 辅助区行范围

    cSKU, cOP, cGRP, cFAC = (CL(RD_H0 + i) for i in range(4))          # AV AW AX AY
    cDT, cAGE, cBKT, cQTY = (CL(RD_H0 + 4 + i) for i in range(4))      # AZ BA BB BC
    cPRC, cAMT, cFLG = (CL(RD_H0 + 8 + i) for i in range(3))           # BD BE BF
    cRQ, cRA, cKEY = (CL(RD_H0 + 11 + i) for i in range(3))            # BG BH BI

    def A(c):
        return "$%s$%d:$%s$%d" % (c, R1, c, RN)

    banner(ws, 1, 12, "合同冗余分析 · 超期未交付占压（口径对齐「倍优待交付情况」，数据源 = 原始数据表）")
    ws.merge_cells("A2:AT2")
    put(ws, 2, 1, "口径：交付周期 = 基准日 − 交货日期（正数 = 已逾期）；超过「冗余判定阈值」仍未到货 = 冗余。"
                  "冗余数量取原始数据表「未到货总数量」，冗余金额 = 未到货总数量 × 采购单价。"
                  "⑤接单判断需先到「销售预测表」填未来3月预计出货数量。右侧 %s 列起为辅助计算区，请勿删改。" % cSKU,
        font=f_note(9, "1F4E79", False), fillc="EAF3FB", align=AL_L)
    ws.row_dimensions[2].height = 32

    # ---------- 参数区（黄底 = 可调假设） ----------
    put(ws, 3, 1, "冗余判定阈值(天)", font=f_formula(9, True), fillc=F_ASSUME, align=AL_L)
    put(ws, 3, 2, 90, font=f_input(10, True), fillc=F_ASSUME, align=AL_C, numfmt="0")
    put(ws, 4, 1, "基准日(留空=今天)", font=f_formula(9, True), fillc=F_ASSUME, align=AL_L)
    put(ws, 4, 2, None, font=f_input(10, True), fillc=F_ASSUME, align=AL_C, numfmt=DATE_FMT)
    put(ws, 5, 1, "有效基准日(自动)", font=f_note(9, "808080", False), align=AL_L)
    put(ws, 5, 2, '=IF($B$4="",TODAY(),$B$4)', font=f_formula(9), align=AL_C, numfmt=DATE_FMT)
    put(ws, 6, 1, "在手数据行数(自动)", font=f_note(9, "808080", False), align=AL_L)
    put(ws, 6, 2, '=COUNTIF(%s,"?*")' % A(cSKU), font=f_formula(9), align=AL_C, numfmt="#,##0")
    put(ws, 7, 1, "阈值说明", font=f_note(9, "808080", False), align=AL_L)
    put(ws, 7, 2, "交付周期 > 阈值 即判为冗余", font=f_note(9, "C55A11", False), align=AL_L)

    # ---------- KPI（被 01驾驶舱⑦ 引用 E3..E9）----------
    K0 = 35   # ⑤区起始列（AI）；④按产品编号块 hd=8 列 → 合并到 34，故 ⑤ 从 35 起避免撞格
    kpis = [
        ("在手待交付总量(件)", "=SUM(%s)" % A(cQTY), "#,##0"),
        ("在手待交付金额(元)", "=SUM(%s)" % A(cAMT), "#,##0"),
        ("冗余数量(超阈值,件)", "=SUM(%s)" % A(cRQ), "#,##0"),
        ("冗余金额(元)", "=SUM(%s)" % A(cRA), "#,##0"),
        ("冗余金额占比", "=IFERROR($E$6/$E$4,0)", "0.0%"),
        ("冗余行数", '=COUNTIF(%s,"冗余")' % A(cFLG), "#,##0"),
        ("不可接单组合数", '=COUNTIF($%s$%d:$%s$%d,"不可接单")'
            % (CL(K0 + 12), RD_R0, CL(K0 + 12), RD_R0 + max(len(dims["combos"]), 1) - 1), "#,##0"),
    ]
    for i, (nm, f, fmt) in enumerate(kpis):
        put(ws, 3 + i, 4, nm, font=Font(name=FONT_BASE, size=9, color="FFFFFF"), fillc="9C0006", align=AL_L)
        put(ws, 3 + i, 5, f, font=Font(name=FONT_BASE, size=13, bold=True, color="9C0006"),
            fillc=F_KPI, align=AL_C, numfmt=fmt)

    # ================= ① 交付周期分档 =================
    ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=6)
    put(ws, 9, 1, "① 交付周期分档（全量在手）", font=f_title(11), fillc=F_KPI, align=AL_L)
    for j, t in enumerate(["交付周期", "行数", "未到货数量", "未到货金额", "金额占比", "风险等级 / 处理建议"], start=1):
        put(ws, 10, j, t, font=f_head(9), fillc=F_SUBHEAD)
    b0 = RD_R0
    btot = b0 + len(RD_BUCKETS)
    r = b0
    for nm, lvl, adv in RD_BUCKETS:
        put(ws, r, 1, nm, font=f_input(9))
        put(ws, r, 2, "=COUNTIF(%s,$A%d)" % (A(cBKT), r), font=f_formula(9), numfmt="#,##0")
        put(ws, r, 3, "=SUMIF(%s,$A%d,%s)" % (A(cBKT), r, A(cQTY)), font=f_formula(9), numfmt="#,##0")
        put(ws, r, 4, "=SUMIF(%s,$A%d,%s)" % (A(cBKT), r, A(cAMT)), font=f_formula(9), numfmt="#,##0")
        put(ws, r, 5, "=IFERROR($D%d/$D$%d,0)" % (r, btot), font=f_formula(9), numfmt="0.0%")
        put(ws, r, 6, "%s ｜ %s" % (lvl, adv), font=f_note(9, "808080", False), align=AL_L)
        r += 1
    put(ws, r, 1, "合计", font=f_formula(9, True), fillc=F_GREY)
    for j in (2, 3, 4):
        put(ws, r, j, "=SUM(%s%d:%s%d)" % (CL(j), b0, CL(j), r - 1),
            font=f_formula(9, True), fillc=F_GREY, numfmt="#,##0")
    put(ws, r, 5, "=IFERROR($D%d/$D%d,0)" % (r, r), font=f_formula(9, True), fillc=F_GREY, numfmt="0.0%")
    put(ws, r, 6, None, fillc=F_GREY)
    ws.conditional_formatting.add("E%d:E%d" % (b0, btot - 1),
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="F8696B"))

    # ================= ② 按运营组别 =================
    G0 = 8
    _dim_block(ws, G0, "② 按运营组别（冗余归属 + 跟进闭环）", dims["groups"], A(cGRP),
               A(cQTY), A(cAMT), A(cRQ), A(cRA), first_hd="运营组别", followup=True)

    # ================= ③ 按厂商 =================
    V0 = 19
    _dim_block(ws, V0, "③ 按厂商（冗余责任方）", dims["facs"], A(cFAC),
               A(cQTY), A(cAMT), A(cRQ), A(cRA), first_hd="厂商简称", followup=False)

    # ================= ④ 按产品编号 =================
    S0 = 27
    _dim_block(ws, S0, "④ 按产品编号（公司级冗余，已按金额降序）", dims["skus"], A(cSKU),
               A(cQTY), A(cAMT), A(cRQ), A(cRA), first_hd="产品编号", followup=False)

    # ================= ⑤ SKU × 运营专员 接单判断 =================
    ws.merge_cells(start_row=9, start_column=K0, end_row=9, end_column=K0 + 12)
    put(ws, 9, K0, "⑤ 接单判断（产品编号 × 运营专员）——「公司级冗余 > 安全库存」即停止接单",
        font=f_title(11), fillc="FCE4D6", align=AL_L)
    hd5 = ["产品编号", "运营专员", "组合键", "待交付数量", "待交付金额", "本组合冗余数量",
           "未来3月预计出货", "公司级冗余(该SKU)", "差额(待交付-预测)", "占公司级冗余比",
           "等级", "月均安全库存", "接单判断"]
    for j, t in enumerate(hd5):
        put(ws, 10, K0 + j, t, font=f_head(9), fillc="C55A11" if j in (6, 12) else F_SUBHEAD)
    cQ3 = CL(K0 + 3)                          # AK 待交付数量
    cJ_ = CL(K0 + 2)                          # AJ 组合键
    cN_ = CL(K0 + 6)                          # AN 未来3月预计出货
    cO_ = CL(K0 + 7)                          # AO 公司级冗余
    cR_ = CL(K0 + 10)                         # AR 等级
    cS_ = CL(K0 + 11)                         # AS 月均安全库存
    FK = "%s!$C$4:$C$503" % Q_FCST            # 销售预测表 组合键
    FV = "%s!$D$4:$D$503" % Q_FCST            # 销售预测表 未来3月预计出货
    k0 = RD_R0
    r = k0
    for sku, op in dims["combos"]:
        gk = "$%s%d" % (CL(K0), r)
        ok = "$%s%d" % (CL(K0 + 1), r)
        put(ws, r, K0 + 0, sku, font=f_input(9), align=AL_L)
        put(ws, r, K0 + 1, op, font=f_input(9), align=AL_L)
        put(ws, r, K0 + 2, '=IF(%s="","",%s&"|"&%s)' % (gk, gk, ok),
            font=f_formula(9), fillc=F_GREY, align=AL_L)
        put(ws, r, K0 + 3, '=IF(%s="","",SUMIFS(%s,%s,%s,%s,%s))'
            % (gk, A(cQTY), A(cSKU), gk, A(cOP), ok), font=f_formula(9), numfmt="#,##0")
        put(ws, r, K0 + 4, '=IF(%s="","",SUMIFS(%s,%s,%s,%s,%s))'
            % (gk, A(cAMT), A(cSKU), gk, A(cOP), ok), font=f_formula(9), numfmt="#,##0")
        put(ws, r, K0 + 5, '=IF(%s="","",SUMIFS(%s,%s,%s,%s,%s))'
            % (gk, A(cRQ), A(cSKU), gk, A(cOP), ok), font=f_formula(9), numfmt="#,##0")
        put(ws, r, K0 + 6, '=IF(%s="","",SUMIF(%s,$%s%d,%s))' % (gk, FK, cJ_, r, FV),
            font=f_xref(9), numfmt="#,##0")
        put(ws, r, K0 + 7, '=IF(%s="","",SUMIF(%s,%s,%s))' % (gk, A(cSKU), gk, A(cRQ)),
            font=f_formula(9), numfmt="#,##0")
        put(ws, r, K0 + 8, '=IF(%s="","",$%s%d-$%s%d)' % (gk, cQ3, r, cN_, r),
            font=f_formula(9), numfmt="#,##0")
        put(ws, r, K0 + 9, '=IF(%s="","",IFERROR($%s%d/$%s%d,0))' % (gk, cQ3, r, cO_, r),
            font=f_formula(9), numfmt="0.0%")
        put(ws, r, K0 + 10, ('=IF(%s="","",IF($%s%d=0,"待填预测",IF($%s%d>500,"S",'
                             'IF($%s%d>200,"A",IF($%s%d>50,"B","C")))))'
                             % (gk, cN_, r, cN_, r, cN_, r, cN_, r)),
            font=f_formula(9, True), align=AL_C)
        put(ws, r, K0 + 11, ('=IF($%s%d="","",IF($%s%d="待填预测","",'
                             'IF($%s%d="S",$%s%d*1.5/3,IF($%s%d="A",$%s%d/3,'
                             'IF($%s%d="B",$%s%d*0.8/3,$%s%d*0.5/3)))))'
                             % (cR_, r, cR_, r, cR_, r, cN_, r, cR_, r, cN_, r,
                                cR_, r, cN_, r, cN_, r)),
            font=f_formula(9), numfmt="#,##0.0")
        put(ws, r, K0 + 12, ('=IF(%s="","",IF($%s%d="待填预测","待填预测",'
                             'IF($%s%d>$%s%d,"不可接单","可接单")))'
                             % (gk, cR_, r, cO_, r, cS_, r)),
            font=f_formula(9, True), align=AL_C)
        r += 1
    k_end = max(r - 1, k0)

    cJUD = CL(K0 + 12)
    rng5 = "%s%d:%s%d" % (cJUD, k0, cJUD, k_end)
    for label, bg, fg, bold in (("不可接单", "FFC7CE", "9C0006", True),
                                ("可接单", "C6EFCE", "1F6E43", False),
                                ("待填预测", "FFF2CC", "9C6500", False)):
        ws.conditional_formatting.add(rng5, FormulaRule(
            formula=['EXACT($%s%d,"%s")' % (cJUD, k0, label)], fill=fill(bg),
            font=Font(name=FONT_BASE, size=9, bold=bold, color=fg), stopIfTrue=False))
    ws.conditional_formatting.add("%s%d:%s%d" % (CL(K0 + 4), k0, CL(K0 + 4), k_end),
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="FFB628"))

    # ================= ⑥ 辅助计算区 =================
    ws.merge_cells(start_row=9, start_column=RD_H0, end_row=9, end_column=RD_H0 + 13)
    put(ws, 9, RD_H0, "⑥ 辅助计算区（逐行直引「原始数据表」，上方所有汇总均取自这里 —— 请勿删除或改动）",
        font=f_title(11), fillc="F2F2F2", align=AL_L)
    hd6 = ["产品编号", "运营专员", "运营组别", "厂商简称", "交货日期", "交付周期(天)", "分档",
           "未到货总数量", "采购单价", "未到货金额", "是否冗余", "冗余数量", "冗余金额", "SKU|专员"]
    for j, t in enumerate(hd6):
        put(ws, 10, RD_H0 + j, t, font=f_head(9), fillc="808080")

    for k in range(RD_NB):
        r = RD_R0 + k
        sr = 3 + k                              # 原始数据表数据行（从第3行起）
        gk = "$%s%d" % (cSKU, r)
        put(ws, r, RD_H0 + 0, '=IF(%s!$C%d="","",%s!$C%d)' % (Q_RAW, sr, Q_RAW, sr),
            font=f_xref(8), align=AL_L)
        put(ws, r, RD_H0 + 1, '=IF(%s="","",%s!$E%d)' % (gk, Q_RAW, sr), font=f_xref(8), align=AL_L)
        put(ws, r, RD_H0 + 2, '=IF(%s="","",%s!$BO%d)' % (gk, Q_RAW, sr), font=f_xref(8), align=AL_L)
        put(ws, r, RD_H0 + 3, '=IF(%s="","",%s!$N%d)' % (gk, Q_RAW, sr), font=f_xref(8), align=AL_L)
        put(ws, r, RD_H0 + 4, '=IF(%s="","",%s!$G%d)' % (gk, Q_RAW, sr), font=f_xref(8), numfmt=DATE_FMT)
        put(ws, r, RD_H0 + 5, '=IF(%s="","",IF($%s%d="","",$B$5-$%s%d))' % (gk, cDT, r, cDT, r),
            font=f_formula(8), numfmt="0")
        put(ws, r, RD_H0 + 6, ('=IF($%s%d="","",IF($%s%d<0,"未到期",IF($%s%d<=90,"逾期1-90",'
                               'IF($%s%d<=180,"逾期91-180",IF($%s%d<=270,"逾期181-270",'
                               'IF($%s%d<=365,"逾期271-365","逾期365以上"))))))'
                               % (cAGE, r, cAGE, r, cAGE, r, cAGE, r, cAGE, r, cAGE, r)),
            font=f_formula(8))
        put(ws, r, RD_H0 + 7, '=IF(%s="","",N(%s!$AG%d))' % (gk, Q_RAW, sr),
            font=f_xref(8), numfmt="#,##0")
        put(ws, r, RD_H0 + 8, '=IF(%s="","",N(%s!$AB%d))' % (gk, Q_RAW, sr),
            font=f_xref(8), numfmt="#,##0.00")
        put(ws, r, RD_H0 + 9, '=IF(%s="","",$%s%d*$%s%d)' % (gk, cQTY, r, cPRC, r),
            font=f_formula(8), numfmt="#,##0")
        put(ws, r, RD_H0 + 10, ('=IF(%s="","",IF($%s%d="","无交期",IF($%s%d>$B$3,"冗余","正常")))'
                                % (gk, cAGE, r, cAGE, r)), font=f_formula(8))
        put(ws, r, RD_H0 + 11, '=IF(%s="","",IF($%s%d="冗余",$%s%d,0))' % (gk, cFLG, r, cQTY, r),
            font=f_formula(8), numfmt="#,##0")
        put(ws, r, RD_H0 + 12, '=IF(%s="","",IF($%s%d="冗余",$%s%d,0))' % (gk, cFLG, r, cAMT, r),
            font=f_formula(8), numfmt="#,##0")
        put(ws, r, RD_H0 + 13, '=IF(%s="","",%s&"|"&$%s%d)' % (gk, gk, cOP, r),
            font=f_formula(8), align=AL_L)

    ws.conditional_formatting.add("%s%d:%s%d" % (cFLG, R1, cFLG, RN), FormulaRule(
        formula=['EXACT($%s%d,"冗余")' % (cFLG, R1)], fill=fill("FFC7CE"),
        font=Font(name=FONT_BASE, size=8, bold=True, color="9C0006"), stopIfTrue=False))

    # ---------- 列宽 ----------
    ws.column_dimensions["A"].width = 15
    for j in range(2, 6):
        ws.column_dimensions[CL(j)].width = 13
    ws.column_dimensions["E"].width = 18          # KPI 金额列加宽，避免大数显示 ########
    ws.column_dimensions["F"].width = 40
    for gap in (7, 18, 26, 33, 47):
        ws.column_dimensions[CL(gap)].width = 3
    for j, w in enumerate([26, 8, 13, 14, 11, 12, 13, 11, 12, 18]):
        ws.column_dimensions[CL(G0 + j)].width = w
    for j, w in enumerate([16, 8, 13, 14, 12, 13, 11, 11]):
        ws.column_dimensions[CL(V0 + j)].width = w
    for j, w in enumerate([22, 8, 13, 14, 12, 13, 11, 11]):
        ws.column_dimensions[CL(S0 + j)].width = w
    for j, w in enumerate([22, 11, 26, 12, 13, 13, 14, 15, 15, 13, 8, 12, 12]):
        ws.column_dimensions[CL(K0 + j)].width = w
    for j, w in enumerate([20, 10, 20, 14, 11, 11, 12, 12, 10, 12, 9, 10, 11, 26]):
        ws.column_dimensions[CL(RD_H0 + j)].width = w
    ws.row_dimensions[9].height = 22
    ws.row_dimensions[10].height = 30
    ws.freeze_panes = "A11"
    ws.sheet_view.showGridLines = False
    return dict(k0=k0, k_end=k_end)


def _dim_block(ws, c0, title, keys, key_rng, qty_rng, amt_rng, rq_rng, ra_rng,
               first_hd="维度", followup=False):
    """通用维度汇总块。followup=True 时额外加两列手工跟进（完成时间 / 结果确认）。"""
    hd = [first_hd, "行数", "待交付数量", "待交付金额", "金额占比", "冗余数量", "冗余金额", "冗余金额率"]
    if followup:
        hd += ["完成时间", "结果确认"]
    ws.merge_cells(start_row=9, start_column=c0, end_row=9, end_column=c0 + len(hd) - 1)
    put(ws, 9, c0, title, font=f_title(11), fillc=F_KPI, align=AL_L)
    for j, t in enumerate(hd):
        put(ws, 10, c0 + j, t, font=f_head(9), fillc="C55A11" if j >= 8 else F_SUBHEAD)

    cKEY = CL(c0)
    cAMT = CL(c0 + 3)
    cRA = CL(c0 + 6)
    r0 = RD_R0
    rtot = r0 + len(keys)
    r = r0
    for k in keys:
        kr = "$%s%d" % (cKEY, r)
        put(ws, r, c0 + 0, k, font=f_input(9), align=AL_L)
        put(ws, r, c0 + 1, "=COUNTIF(%s,%s)" % (key_rng, kr), font=f_formula(9), numfmt="#,##0")
        put(ws, r, c0 + 2, "=SUMIF(%s,%s,%s)" % (key_rng, kr, qty_rng), font=f_formula(9), numfmt="#,##0")
        put(ws, r, c0 + 3, "=SUMIF(%s,%s,%s)" % (key_rng, kr, amt_rng), font=f_formula(9), numfmt="#,##0")
        put(ws, r, c0 + 4, "=IFERROR($%s%d/$%s$%d,0)" % (cAMT, r, cAMT, rtot), font=f_formula(9), numfmt="0.0%")
        put(ws, r, c0 + 5, "=SUMIF(%s,%s,%s)" % (key_rng, kr, rq_rng), font=f_formula(9), numfmt="#,##0")
        put(ws, r, c0 + 6, "=SUMIF(%s,%s,%s)" % (key_rng, kr, ra_rng), font=f_formula(9), numfmt="#,##0")
        put(ws, r, c0 + 7, "=IFERROR($%s%d/$%s%d,0)" % (cRA, r, cAMT, r), font=f_formula(9, True), numfmt="0.0%")
        if followup:
            put(ws, r, c0 + 8, None, font=f_input(9), fillc="FFF0E1", numfmt=DATE_FMT)
            put(ws, r, c0 + 9, None, font=f_input(9), fillc="FFF0E1", align=AL_L)
        r += 1
    put(ws, r, c0, "合计", font=f_formula(9, True), fillc=F_GREY)
    for j in (1, 2, 3, 5, 6):
        put(ws, r, c0 + j, "=SUM(%s%d:%s%d)" % (CL(c0 + j), r0, CL(c0 + j), r - 1),
            font=f_formula(9, True), fillc=F_GREY, numfmt="#,##0")
    put(ws, r, c0 + 4, "=IFERROR($%s%d/$%s%d,0)" % (cAMT, r, cAMT, r),
        font=f_formula(9, True), fillc=F_GREY, numfmt="0.0%")
    put(ws, r, c0 + 7, "=IFERROR($%s%d/$%s%d,0)" % (cRA, r, cAMT, r),
        font=f_formula(9, True), fillc=F_GREY, numfmt="0.0%")
    if followup:
        put(ws, r, c0 + 8, None, fillc=F_GREY)
        put(ws, r, c0 + 9, None, fillc=F_GREY)
    if keys:
        ws.conditional_formatting.add("%s%d:%s%d" % (cRA, r0, cRA, rtot - 1),
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="F8696B"))
        ws.conditional_formatting.add("%s%d:%s%d" % (cAMT, r0, cAMT, rtot - 1),
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="638EC6"))
    return rtot
