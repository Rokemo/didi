# -*- coding: utf-8 -*-
import os, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter as CL
from build_part1 import *
from build_util import *
import build_sheets_a as A
import build_sheets_b as B
import build_sheets_c as C
import build_sheets_d as D

OUT = r"C:/Users/beiyou201/WorkBuddy/2026-07-31-18-27-53/理单-跟单进度跟踪表_v6.xlsx"


def load_raw():
    """读取用户维护的原始数据.xlsx（WPS私有格式），规整为跟单订单结构"""
    rows = wps_read(NEW_FILE, 1)
    set_manual_cols(len(rows[0]))          # 依据实际导出列数定位 4 个手工维护列（紧贴导出列之后）
    orders, mats, facs, buyers = norm_raw(rows)
    return rows, orders, mats, facs, buyers


def build_readme(wb, stats):
    ws = wb.create_sheet(SH_README)
    banner(ws, 1, 8, "跨境采购跟单进度表 v4 · 使用说明")
    r = 3

    def h(t):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        put(ws, r, 1, t, font=f_title(12), fillc=F_KPI, align=AL_L)
        ws.row_dimensions[r].height = 24
        r += 1

    def p(t, bold=False, color="333333", indent=0):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(row=r, column=1, value=t)
        c.font = Font(name=FONT_BASE, size=10, bold=bold, color=color)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=indent)
        ws.row_dimensions[r].height = 18 if len(t) < 60 else 30
        r += 1

    def tbl(headers, rows_, widths):
        nonlocal r
        for j, t in enumerate(headers, start=1):
            put(ws, r, j, t, font=f_head(9), fillc=F_SUBHEAD)
        r += 1
        for row in rows_:
            for j, v in enumerate(row, start=1):
                put(ws, r, j, v, font=f_formula(9) if j > 1 else f_formula(9, True), align=AL_L)
            ws.row_dimensions[r].height = 20
            r += 1

    h("一、这份表解决什么问题")
    p("原表把「订单清单」和「工序节拍模板」放在两个互不相干的地方 —— 看总表只知道欠多少件，"
      "不知道卡在哪道工序；看工序表只有理论排期，不知道实际做到哪一步。v3 把两者打通，"
      "并以「原始数据表 / 产品资料表 / 发货需求表 / 销售预测表」四张为主数据源：你只维护这四张，其余 12 张全部自动联动。")
    p("现在你可以直接回答三个问题：① 这个合同能不能整单出？② 这个 SKU 现在卡在哪道工序？③ 今天该催谁？", bold=True, color="1F4E79")
    r += 1

    h("二、数据怎么流转（v4 核心改动）")
    p("① 四张「你维护」的主数据源（其余表全部引用它们自动联动）：\n"
      "   · 原始数据表：ERP 导出整表粘贴到左侧，右侧 4 列（最新跟进日期 / 跟进结论 / 跟进备注 / 是否新单）手填；\n"
      "   · 产品资料表：产品编号 → 工艺品类。类目不再由系统推断，02 按产品编号 VLOOKUP 取工艺品类，留空=待确认；\n"
      "   · 发货需求表：每行一条发货需求（SKU/团队/发货数量/发货时间），供「10发货需求判断」汇总对比；\n"
      "   · 销售预测表：每行一个「产品编号 × 运营专员」组合（已预生成），你只需填 D 列「未来3月预计出货数量」，"
      "供「11合同冗余」的接单判断（S/A/B/C 分级 + 安全库存）使用。", bold=True, color="1F4E79")
    p("② 02订单明细：不再手填 —— 它的每一格（工厂/合同号/SKU/数量/日期/跟进…）都是绿色公式，实时引用「原始数据表」同一行号。你在原始数据表改一行，02 立刻变。", color="1F6E43")
    p("③ 03~06 及驾驶舱：全部引用 02，环环相扣，无需手动刷新。", color="808080")
    p("⚠ 不要在 02 手工覆盖绿字单元格；新数据请在「原始数据表」里增行，02 会按行号自动对应。", bold=True, color="C00000")
    r += 1

    h("三、每天怎么用（3 分钟例行）")
    tbl(["步骤", "去哪张表", "做什么"],
        [["1", "01驾驶舱", "扫一眼 KPI：逾期行数 / 呆滞行数 / 紧急行数有没有变化"],
         ["2", "原始数据表", "把 ERP 新导出的行粘贴到左侧，或更新右侧手工维护列"],
         ["3", "05逾期呆滞专项", "从上往下看「仍逾期」的行，填「催办动作」和「工厂承诺交期」"],
         ["4", "03工序进度", "把工厂当天报的工序完成日期，填进橙色的「★实际完成」列"],
         ["5", "06合同级汇总", "看哪些合同「缺口件数」已经归零 → 可以约柜订舱了"]],
        None)
    r += 1

    h("四、颜色约定（行业标准，别改）")
    cs = [("蓝色字 + 浅蓝底", C_INPUT, F_INPUT, "手工录入区。只有这些格子需要你动手"),
          ("蓝色字 + 橙色底", C_INPUT, "FFF0E1", "高频录入区（工序实际完成日 / 催办动作），每天都要填"),
          ("黑色字", C_FORMULA, None, "本表内公式，自动算，别覆盖"),
          ("绿色字", C_XREF, None, "跨表引用（02 全部引用原始数据表），改源头表这里自动变"),
          ("黄色底", "C55A11", F_ASSUME, "假设值 / 待确认项，需要你去核实后修正"),
          ("红底红字", "9C0006", "FFC7CE", "告警：逾期、延期、高风险")]
    for j, t in enumerate(["样式", "含义"], start=1):
        put(ws, r, j, t, font=f_head(9), fillc=F_SUBHEAD)
    r += 1
    for name, fc, bg, desc in cs:
        put(ws, r, 1, name, font=Font(name=FONT_BASE, size=10, bold=True, color=fc), fillc=bg)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        put(ws, r, 2, desc, font=f_formula(9), align=AL_L)
        r += 1
    r += 1

    h("五、16 张表分别干什么")
    tbl(["表名", "性质", "说明"],
        [[SH_RAW, "★你维护", "主数据源之一。左侧粘贴ERP导出，右侧填跟进。其余表都引用它"],
         [SH_PROD, "★你维护", "产品资料表：产品编号 → 工艺品类。类目唯一来源，02 按产品编号 VLOOKUP"],
         [SH_SHIP, "★你维护", "发货需求表：每行一条发货需求（SKU/团队/发货数量/发货时间）"],
         [SH_DASH, "只看不改", "全局 KPI + 按工厂/类目/跟单员/账龄/发货需求的多元汇总，全部公式联动"],
         [SH_ORDER, "自动引用", "一行 = 一个「合同号 + SKU」。绿字来自原始数据表，标准工艺类目来自产品资料表"],
         [SH_PROC, "半自动", "9 道统一工序，计划日期按节拍自动倒排；你只填「实际完成」，卡点自动跳出来"],
         [SH_GANTT, "只看不改", "滚动 22 周甘特，颜色深浅 = 计划推进度，红块 = 已过交期"],
         [SH_LATE, "半自动", "逾期清单，按逾期天数降序。右侧 4 列橙色是你的催办台账"],
         [SH_CONT, "只看不改", "按母合同号归集（母合同号=合同号全文），回答「能不能整单出货」"],
         [SH_TACT, "参数", "工序节拍库。改这里 = 全表排期重算。黄底行是回落值，要跟工厂确认"],
         [SH_DICT, "参数", "标准工艺类目总表、工序名对照、预警阈值、下拉选项源"],
         [SH_QC, "参考", "产中检验时间标准（原 Sheet1），工序进「进行中」时对照安排验货"],
         [SH_SHIPCHK, "半自动", "发货需求判断：设发货起止日，按产品编号+团队对比未入库数量，看能否满足"],
         [SH_FCST, "★你维护", "销售预测表：产品编号×运营专员 → 未来3月预计出货数量（黄底D列）。接单判断的唯一数据来源，填后 11合同冗余 才生效"],
         [SH_REDUN, "只看不改", "合同冗余分析：超期未交付占压 + SKU×运营专员接单判断。数据源=原始数据表，口径对齐「倍优待交付情况」"]],
        None)
    r += 1

    h("六、导入这份数据时发现的问题（请优先处理）")
    warns = [
        ("🔴 高", "产品资料表「工艺品类」尚未填写 → 类目全待确认",
         "本次导入的产品资料表 1675 行的「工艺品类」列为空。类目不再由系统推断，需你在「产品资料表」"
         "按产品编号补全工艺品类（可下拉选标准类目），02/03 才会出排期，否则整列显示「待确认」。"),
        ("🔴 高", "大量行逾期，其中多行逾期超过 180 天",
         "最久的已挂账 800+ 天。这类尾数订单挂在表上没有意义，建议单独开清尾评审（核销 / 退单 / 转内销）。"),
        ("🟠 中", "发货需求与原始数据的「团队」口径需对齐",
         "发货需求判断按 产品编号+团队 汇总，团队字符串须与原始数据表完全一致（如 鄢紫薇/UP-US）。"
         "不一致会导致可供应未入库=0、误报缺口；请核对两边团队写法。"),
        ("🟠 中", "节拍参数缺若干组合",
         "原模板只给了部分「数量档位×新单」组合。07参数-工序节拍 中黄底行是按规则回落填的，"
         "请工厂确认后覆盖。"),
        ("🟡 低", "合并单元格导致字段空白",
         "原总表靠合并单元格视觉填充工厂/合同号，无法筛选透视。v4 已逐行补齐。"),
        ("🟠 中", "销售预测表 D 列为空 → 接单判断全部显示「待填预测」",
         "「11合同冗余」⑤区需要一个未来3月预计出货数量（来自「销售预测表」D列）才能算 S/A/B/C 等级与安全库存。"
         "当前该列为黄底空值，⑤区会显示「待填预测」、不报可/不可接单。请按实际销售预期填写后重新查看。"),
        ("🟠 中", "大量行交货日期在未来 → 交付周期分档以「未到期」为主",
         "原始数据 500 行中约 401 行交货日期晚于今天，交付周期=基准日−交货日期 为负，判为「未到期」（在途正常单）。"
         "但已有 25 行超过 90 天阈值被判为「冗余」（冗余数量 262 件 / 金额约 14.3 万元），需优先催办；"
         "待更多订单实际逾期后，①分档与⑤接单判断会逐步体现占压。"),
    ]
    for j, t in enumerate(["等级", "问题", "说明与建议"], start=1):
        put(ws, r, j, t, font=f_head(9), fillc="9C0006")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    r += 1
    for lv, tt, dd in warns:
        put(ws, r, 1, lv, font=f_formula(9, True))
        put(ws, r, 2, tt, font=f_formula(9, True), align=AL_L)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        put(ws, r, 3, dd, font=f_formula(9), align=AL_L)
        ws.row_dimensions[r].height = 44
        r += 1
    r += 1

    h("七、新增 / 更新订单怎么操作（v3 最简路径）")
    p("1) 打开「原始数据表」：把 ERP 新导出的整行粘贴到左侧 85 列（行号与 02 自动对齐），右侧 4 列填跟进。")
    p("2) 是否新单：在右侧「是否新单」列选 是/否（影响节拍与预警）。")
    p("3) 其余 02 / 03 / 04 / 05 / 06 / 驾驶舱 全部自动联动，无需任何手动下拉公式。", bold=True, color="1F6E43")
    p("4) 新工厂 / 新 SKU：标准工艺类目请在「产品资料表」按产品编号补全「工艺品类」（可下拉选）；"
      "新工厂在原始数据表追加即可，下拉选项源会自动更新。")
    p("5) 发货需求判断：在「发货需求表」增删发货需求行（SKU/团队/发货数量/发货时间），"
      "到「10发货需求判断」设发货起止日，即可看窗口内需求能否被未入库数量满足、缺口多少。")
    p("6) 合同冗余与接单判断：先在「销售预测表」D 列填各组合未来3月预计出货数量，"
      "再到「11合同冗余」看超期占压（按运营组别/厂商/SKU）与「公司级冗余 > 安全库存」的不可接单组合。")
    r += 1

    h("八、9 道统一工序是怎么来的")
    p("原表 6 张类目模板工序数量不一致（树脂 2 道、普通铁艺 6 道、转印 8 道）。"
      "v3 取并集抽象成 9 道通用工序框架，各类目用本地叫法映射，不适用的显示「—」并自动跳过。")
    tbl(["统一工序", "特斯林", "编藤", "铁艺铝艺-普通", "树脂"],
        [["P%d %s" % (i+1, STAGES[i]),
          CAT_STAGE_NAME["特斯林"][i], CAT_STAGE_NAME["编藤"][i],
          CAT_STAGE_NAME["铁艺铝艺-普通"][i], CAT_STAGE_NAME["树脂"][i]] for i in range(9)],
        None)
    p("完整 6 类目对照见 08参数-字典 ②区。", color="808080")

    setw(ws, [22, 30, 20, 16, 16, 16, 16, 16])
    ws.column_dimensions["C"].width = 24
    ws.sheet_view.showGridLines = False


def main():
    raw_rows, orders, mats, facs, buyers = load_raw()
    dims = redun_dims(raw_rows)
    prod_rows = load_prod()
    ship_rows = load_ship()
    qc = load_qc()
    wb = Workbook()
    wb.remove(wb.active)

    tact_last = A.build_tact(wb)
    dinfo = A.build_dict(wb, mats, facs, buyers)
    A.build_raw(wb, raw_rows)
    A.build_prod(wb, prod_rows)
    A.build_ship(wb, ship_rows)
    A.build_fcst(wb, dims["combos"], {})        # 销售预测表（用户填 D 列）
    last = B.build_order(wb, orders, dinfo)
    B.build_proc(wb, orders, dinfo, tact_last)
    C.build_contract(wb, orders, last)
    C.build_dash(wb, orders, last, dinfo)
    C.build_gantt(wb, last)
    C.build_late(wb, orders, dinfo)
    A.build_qc(wb, qc)
    C.build_shipchk(wb, len(ship_rows))
    D.build_redun(wb, dims)                       # 11合同冗余（数据源=原始数据表）
    build_readme(wb, {})

    order = [SH_README, SH_DASH, SH_RAW, SH_PROD, SH_SHIP, SH_ORDER, SH_PROC, SH_GANTT,
             SH_LATE, SH_CONT, SH_TACT, SH_DICT, SH_QC, SH_SHIPCHK, SH_FCST, SH_REDUN]
    wb._sheets = [wb[n] for n in order]
    wb.active = 0
    for n in order:
        wb[n].sheet_properties.tabColor = {
            SH_README: "808080", SH_DASH: "1F4E79", SH_RAW: "C55A11", SH_ORDER: "C55A11",
            SH_PROC: "1F6E43", SH_GANTT: "5B3A8E", SH_LATE: "9C0006", SH_CONT: "2E75B6",
            SH_TACT: "BF8F00", SH_DICT: "BF8F00", SH_QC: "7F7F7F",
            SH_PROD: "2E75B6", SH_SHIP: "C55A11", SH_SHIPCHK: "548235",
            SH_FCST: "548235", SH_REDUN: "9C0006"}[n]
    wb.save(OUT)
    print("saved:", OUT, os.path.getsize(OUT), "bytes")
    print("sheets:", wb.sheetnames)
    print("订单行:", len(orders), "合同数:", len({o['con'] for o in orders}),
          "工厂:", len(facs), "跟单员:", len(buyers),
          "产品资料:", len(prod_rows), "发货需求:", len(ship_rows))


if __name__ == "__main__":
    main()
