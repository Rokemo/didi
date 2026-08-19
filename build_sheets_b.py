# -*- coding: utf-8 -*-
"""核心表：02订单明细 / 03工序进度"""
from openpyxl.utils import get_column_letter as CL
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule, ColorScaleRule
from build_part1 import *
from build_util import *

DATE_FMT = "yyyy-mm-dd"
R0 = 3          # 数据起始行（与原始数据表同行号对齐：02行r ↔ 原始表行r）
Q_TACT = q(SH_TACT)
Q_DICT = q(SH_DICT)
Q_ORD  = q(SH_ORDER)
Q_PRC  = q(SH_PROC)
Q_RAW  = q(SH_RAW)
Q_PROD = q(SH_PROD)

ORD_HEAD = [
    ("序号", 6, "f"), ("工厂", 9, "x"), ("合同号", 25, "x"),
    ("SKU", 21, "x"), ("标准工艺类目", 14, "x"),
    ("订单总数量", 11, "x"), ("已交付数量", 11, "x"), ("待交付数量", 11, "f"), ("完成率", 10, "f"),
    ("合同总数", 11, "f"), ("数量档位", 11, "f"), ("是否新单", 9, "x"), ("跟单员", 10, "x"),
    ("合同日期", 12, "x"), ("计划出货日", 12, "x"), ("生产周期(天)", 11, "f"),
    ("距出货(天)", 11, "f"), ("交期状态", 11, "f"), ("逾期天数", 10, "f"), ("风险等级", 10, "f"),
    ("工序进度", 10, "x"), ("当前在制工序", 15, "x"), ("卡点状态", 12, "x"),
    ("最新跟进日期", 13, "x"), ("跟进结论", 18, "x"), ("跟进备注", 40, "x"),
]
ORD_GROUPS = [("一、基础订单信息", 1, 5), ("二、交付数量", 6, 11), ("三、订单属性", 12, 16),
              ("四、交期管控", 17, 20), ("五、生产进度（自动取自 03工序进度）", 21, 23),
              ("六、跟单动作（取自原始数据表）", 24, 26)]
GROUP_COLORS = ["1F4E79", "1F6E43", "7F6000", "9C0006", "5B3A8E", "3B3838"]


def build_order(wb, orders, dinfo):
    ws = wb.create_sheet(SH_ORDER)
    n = len(orders)
    last = R0 + n - 1
    thr = dinfo["thr_start"]
    T_URG = "%s!$B$%d" % (Q_DICT, thr)
    T_WRN = "%s!$B$%d" % (Q_DICT, thr+1)
    T_DEAD = "%s!$B$%d" % (Q_DICT, thr+2)
    T_LR1 = "%s!$B$%d" % (Q_DICT, thr+3)
    T_LR2 = "%s!$B$%d" % (Q_DICT, thr+4)

    for i, (name, c1, c2) in enumerate(ORD_GROUPS):
        ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        c = ws.cell(row=1, column=c1, value=name)
        c.font = Font(name=FONT_BASE, size=10, bold=True, color="FFFFFF")
        c.fill = fill(GROUP_COLORS[i]); c.alignment = AL_C; c.border = BOX
        for cc in range(c1+1, c2+1):
            ws.cell(row=1, column=cc).fill = fill(GROUP_COLORS[i])
    ws.row_dimensions[1].height = 22

    for j, (name, w, kind) in enumerate(ORD_HEAD, start=1):
        put(ws, 2, j, name, font=f_head(), fillc=F_HEADER)
        ws.column_dimensions[CL(j)].width = w
    ws.row_dimensions[2].height = 32

    for k, o in enumerate(orders):
        r = R0 + k
        # 数据列：全部绿色跨表引用「原始数据表」（同一行号 r）
        put(ws, r, 1, k+1, font=f_formula(9))
        put(ws, r, 2, "=%s!N%d" % (Q_RAW, r), font=f_xref(), align=AL_L)        # 工厂
        put(ws, r, 3, "=%s!AE%d" % (Q_RAW, r), font=f_xref(), align=AL_L)       # 合同号
        put(ws, r, 4, "=%s!C%d" % (Q_RAW, r), font=f_xref(), align=AL_L)        # SKU
        put(ws, r, 5, '=IFERROR(IF(VLOOKUP($E%d,%s!$A$4:$C$2000,3,FALSE)="","待确认",'
            'VLOOKUP($E%d,%s!$A$4:$C$2000,3,FALSE)),"待确认")' % (r, Q_PROD, r, Q_PROD),
            font=f_xref())    # 标准工艺类目（来自产品资料表 产品编号→工艺品类）
        put(ws, r, 6, "=%s!AD%d" % (Q_RAW, r), font=f_xref(), numfmt="#,##0")   # 订单总数量
        put(ws, r, 7, "=%s!AO%d" % (Q_RAW, r), font=f_xref(), numfmt="#,##0")   # 已交付数量
        put(ws, r, 8, '=SUMIFS(%s!$AG$3:$AG$602,%s!$AE$3:$AE$602,$C%d,%s!$C$3:$C$602,$D%d)'
            % (Q_RAW, Q_RAW, r, Q_RAW, r), font=f_xref(), numfmt="#,##0")           # 待交付数量=SUMIFS(原始表AG 未到货总数量, 合同号+SKU)
        put(ws, r, 9, "=IFERROR(G%d/F%d,0)" % (r, r), font=f_formula(), numfmt="0.0%")  # 完成率=已交付/订单总数量
        put(ws, r, 10, '=SUMIFS($F$%d:$F$%d,$C$%d:$C$%d,$C%d)' % (R0, last, R0, last, r),
            font=f_formula(bold=True), numfmt="#,##0")                           # 合同总数=按合同号汇总订单总数量
        put(ws, r, 11, '=IF(J%d<=1000,"1000以内",IF(J%d<=3000,"1000-3000","3000以上"))' % (r, r), font=f_formula(9))  # 数量档位=按合同总数判定
        put(ws, r, 12, "=%s!%s%d" % (Q_RAW, CL(MANUAL_COLS["是否新单"]), r), font=f_xref())  # 是否新单(动态列)
        put(ws, r, 13, "=%s!I%d" % (Q_RAW, r), font=f_xref())                   # 跟单员
        put(ws, r, 14, "=%s!F%d" % (Q_RAW, r), font=f_xref(), numfmt=DATE_FMT)  # 合同日期
        put(ws, r, 15, "=%s!G%d" % (Q_RAW, r), font=f_xref(), numfmt=DATE_FMT)  # 计划出货日(交货日期)
        put(ws, r, 16, "=O%d-N%d" % (r, r), font=f_formula(9), numfmt="0")      # 生产周期=计划出货日-合同日期
        put(ws, r, 17, "=O%d-TODAY()" % r, font=f_formula(9), numfmt="0")       # 距出货=计划出货日-TODAY
        put(ws, r, 18, '=IF(H%d<=0,"已交付",IF(Q%d<0,"已逾期",IF(Q%d<=%s,"紧急",IF(Q%d<=%s,"预警","正常"))))'
            % (r, r, r, T_URG, r, T_WRN), font=f_formula(bold=True))             # 交期状态
        put(ws, r, 19, "=IF(AND(H%d>0,Q%d<0),-Q%d,0)" % (r, r, r), font=f_formula(9), numfmt="0")  # 逾期天数
        put(ws, r, 20, ('=IF(H%d<=0,"无",IF(S%d>%s,"呆滞",IF(S%d>0,"高",'
                        'IF(AND(Q%d<=%s,K%d<%s),"高",IF(AND(Q%d<=%s,K%d<%s),"中","低")))))')
            % (r, r, T_DEAD, r, r, T_URG, r, T_LR1, r, T_WRN, r, T_LR2), font=f_formula(bold=True))  # 风险等级
        put(ws, r, 21, '=IFERROR(%s!L%d,"")' % (Q_PRC, r), font=f_xref(), numfmt="0%")
        put(ws, r, 22, '=IFERROR(%s!M%d,"")' % (Q_PRC, r), font=f_xref(9))
        put(ws, r, 23, '=IFERROR(%s!N%d,"")' % (Q_PRC, r), font=f_xref(9, bold=True))
        put(ws, r, 24, "=%s!%s%d" % (Q_RAW, CL(MANUAL_COLS["最新跟进日期"]), r), font=f_xref(9), numfmt=DATE_FMT)  # 最新跟进日期(动态列)
        put(ws, r, 25, "=%s!%s%d" % (Q_RAW, CL(MANUAL_COLS["跟进结论"]), r), font=f_xref(9), align=AL_L)     # 跟进结论(动态列)
        put(ws, r, 26, "=%s!%s%d" % (Q_RAW, CL(MANUAL_COLS["跟进备注"]), r), font=f_xref(9), align=AL_L)      # 跟进备注(动态列)

    ws.freeze_panes = "F3"
    ws.auto_filter.ref = "A2:Z%d" % last

    # ---- 条件格式 ----
    rng_t = "R%d:R%d" % (R0, last)
    for txt, bg, fg in [("已逾期", "FFC7CE", "9C0006"), ("紧急", "FFD9B3", "974706"),
                        ("预警", "FFEB9C", "7F6000"), ("正常", "C6EFCE", "1F6E43"),
                        ("已交付", "E7E6E6", "808080")]:
        ws.conditional_formatting.add(rng_t, FormulaRule(
            formula=['EXACT($R%d,"%s")' % (R0, txt)], fill=fill(bg),
            font=Font(name=FONT_BASE, size=10, bold=True, color=fg), stopIfTrue=False))
    rng_v = "T%d:T%d" % (R0, last)
    for txt, bg, fg in [("呆滞", "404040", "FFFFFF"), ("高", "FFC7CE", "9C0006"),
                        ("中", "FFEB9C", "7F6000"), ("低", "C6EFCE", "1F6E43"),
                        ("无", "E7E6E6", "808080")]:
        ws.conditional_formatting.add(rng_v, FormulaRule(
            formula=['EXACT($T%d,"%s")' % (R0, txt)], fill=fill(bg),
            font=Font(name=FONT_BASE, size=10, bold=True, color=fg), stopIfTrue=False))
    ws.conditional_formatting.add("I%d:I%d" % (R0, last),
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="63C384"))
    ws.conditional_formatting.add("U%d:U%d" % (R0, last),
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="8FAADC"))
    ws.conditional_formatting.add("A%d:Z%d" % (R0, last), FormulaRule(
        formula=['$E%d="待确认"' % R0], fill=fill(F_ASSUME), stopIfTrue=False))
    ws.conditional_formatting.add("C%d:E%d" % (R0, last), FormulaRule(
        formula=['$T%d="呆滞"' % R0],
        font=Font(name=FONT_BASE, size=10, color="7F7F7F"), stopIfTrue=False))
    return last


# ================= 03 工序进度 =================
def col_plan(i):   return 15 + (i-1)*3
def col_due(i):    return 16 + (i-1)*3
def col_act(i):    return 17 + (i-1)*3
MIR_PLAN = 42
MIR_DUE = 51
MIR_ACT = 60
MIR_FLAG = 69


def build_proc(wb, orders, dinfo, tact_last):
    ws = wb.create_sheet(SH_PROC)
    n = len(orders)
    last = R0 + n - 1
    TD = "%s!$E$5:$V$%d" % (Q_TACT, tact_last)
    TK = "%s!$D$5:$D$%d" % (Q_TACT, tact_last)
    SD = "%s!$B$%d:$J$%d" % (Q_DICT, dinfo["stage_start"], dinfo["stage_end"])
    SK = "%s!$A$%d:$A$%d" % (Q_DICT, dinfo["stage_start"], dinfo["stage_end"])

    fixed = [("序号", 6), ("工厂", 9), ("合同号", 25), ("SKU", 21), ("标准工艺类目", 14),
             ("数量档位", 10), ("新单", 7), ("合同日期", 12), ("计划出货日", 12),
             ("适用工序数", 10), ("已完成", 9), ("工序进度", 10), ("当前在制工序", 15), ("卡点状态", 12)]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    c = ws.cell(row=1, column=1, value="订单信息（自动取自 02订单明细）")
    c.font = f_head(); c.fill = fill("1F4E79"); c.alignment = AL_C
    for cc in range(2, 10):
        ws.cell(row=1, column=cc).fill = fill("1F4E79")
    ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=14)
    c = ws.cell(row=1, column=10, value="进度汇总（自动）")
    c.font = f_head(); c.fill = fill("5B3A8E"); c.alignment = AL_C
    for cc in range(11, 15):
        ws.cell(row=1, column=cc).fill = fill("5B3A8E")
    for i, s in enumerate(STAGES):
        c1 = col_plan(i+1)
        ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c1+2)
        c = ws.cell(row=1, column=c1, value="P%d  %s" % (i+1, s))
        c.font = f_head(); c.fill = fill("1F6E43" if i % 2 == 0 else "2E7D5B"); c.alignment = AL_C
        for cc in range(c1+1, c1+3):
            ws.cell(row=1, column=cc).fill = fill("1F6E43" if i % 2 == 0 else "2E7D5B")
    ws.row_dimensions[1].height = 24

    for j, (nm, w) in enumerate(fixed, start=1):
        put(ws, 2, j, nm, font=f_head(9), fillc=F_HEADER)
        ws.column_dimensions[CL(j)].width = w
    for i in range(9):
        c1 = col_plan(i+1)
        put(ws, 2, c1, "计划开工", font=f_head(9), fillc="1F6E43")
        put(ws, 2, c1+1, "最晚完工", font=f_head(9), fillc="1F6E43")
        put(ws, 2, c1+2, "★实际完成", font=f_head(9), fillc="C55A11")
        for k in range(3):
            ws.column_dimensions[CL(c1+k)].width = 11
    ws.row_dimensions[2].height = 30
    for base_, nm in [(MIR_PLAN, "开工镜像"), (MIR_DUE, "完工镜像"), (MIR_ACT, "实际镜像"), (MIR_FLAG, "完成标记")]:
        for i in range(9):
            put(ws, 2, base_+i, "%s%d" % (nm, i+1), font=Font(name=FONT_BASE, size=8, color="BFBFBF"), fillc=F_GREY)
            ws.column_dimensions[CL(base_+i)].width = 9
            ws.column_dimensions[CL(base_+i)].hidden = True

    for k in range(n):
        r = R0 + k
        put(ws, r, 1, "=%s!A%d" % (Q_ORD, r), font=f_xref(9))
        # 02列号：SKU=D, 标准工艺类目=E, 数量档位=K, 是否新单=L, 合同日期=N, 计划出货日=O
        for col, src in [(2, "B"), (3, "C"), (4, "D"), (5, "E"), (6, "K"), (7, "L"), (8, "N"), (9, "O")]:
            put(ws, r, col, "=%s!%s%d" % (Q_ORD, src, r), font=f_xref(9),
                align=AL_L if col in (3, 4) else AL_C,
                numfmt=DATE_FMT if col in (8, 9) else None)
        put(ws, r, 10, "=COUNT($%s%d:$%s%d)" % (CL(MIR_PLAN), r, CL(MIR_PLAN+8), r), font=f_formula(9))
        put(ws, r, 11, "=COUNT($%s%d:$%s%d)" % (CL(MIR_ACT), r, CL(MIR_ACT+8), r), font=f_formula(9, True))
        put(ws, r, 12, '=IF($J%d=0,"",$K%d/$J%d)' % (r, r, r), font=f_formula(9), numfmt="0%")
        flag_rng = "$%s$%d:$%s$%d" % (CL(MIR_FLAG), r, CL(MIR_FLAG+8), r)
        due_rng = "$%s$%d:$%s$%d" % (CL(MIR_DUE), r, CL(MIR_DUE+8), r)
        plan_rng = "$%s$%d:$%s$%d" % (CL(MIR_PLAN), r, CL(MIR_PLAN+8), r)
        put(ws, r, 13, ('=IF($J%d=0,"已交付",IFERROR(INDEX(INDEX(%s,MATCH($E%d,%s,0),0),'
                        'MATCH(0,%s,0)),"全部完成"))')
            % (r, SD, r, SK, flag_rng), font=f_formula(9, True))
        put(ws, r, 14, ('=IF($J%d=0,"全部完成",IF($K%d>=$J%d,"全部完成",'
                        'IF(INDEX(%s,MATCH(0,%s,0))<TODAY(),"工序延期",'
                        'IF(INDEX(%s,MATCH(0,%s,0))<=TODAY(),"进行中","未开工"))))')
            % (r, r, r, due_rng, flag_rng, plan_rng, flag_rng), font=f_formula(9, True))

        key = '$E%d&"|"&$F%d&"|"&$G%d' % (r, r, r)
        for i in range(1, 10):
            cp, cd, ca = col_plan(i), col_due(i), col_act(i)
            idx_s, idx_e = (i-1)*2+1, (i-1)*2+2
            put(ws, r, cp, ('=IFERROR(IF(INDEX(%s,MATCH(%s,%s,0),%d)="NA","—",'
                            '$H%d+INDEX(%s,MATCH(%s,%s,0),%d)),"待确认")')
                % (TD, key, TK, idx_s, r, TD, key, TK, idx_s),
                font=f_formula(9), numfmt=DATE_FMT)
            put(ws, r, cd, ('=IFERROR(IF(INDEX(%s,MATCH(%s,%s,0),%d)="NA","—",'
                            '$I%d+INDEX(%s,MATCH(%s,%s,0),%d)),"待确认")')
                % (TD, key, TK, idx_e, r, TD, key, TK, idx_e),
                font=f_formula(9, True), numfmt=DATE_FMT)
            put(ws, r, ca, None, font=f_input(9), fillc="FFF0E1", numfmt=DATE_FMT)
            put(ws, r, MIR_PLAN+i-1, '=IF(ISNUMBER(%s%d),%s%d,"")' % (CL(cp), r, CL(cp), r),
                font=f_formula(8), numfmt=DATE_FMT, border=False)
            put(ws, r, MIR_DUE+i-1, '=IF(ISNUMBER(%s%d),%s%d,"")' % (CL(cd), r, CL(cd), r),
                font=f_formula(8), numfmt=DATE_FMT, border=False)
            put(ws, r, MIR_ACT+i-1, '=IF(ISNUMBER(%s%d),%s%d,"")' % (CL(ca), r, CL(ca), r),
                font=f_formula(8), numfmt=DATE_FMT, border=False)
            put(ws, r, MIR_FLAG+i-1, '=IF(%s%d="",1,IF(%s%d="",0,1))'
                % (CL(MIR_PLAN+i-1), r, CL(MIR_ACT+i-1), r), font=f_formula(8), border=False)

    ws.freeze_panes = "E3"
    ws.auto_filter.ref = "A2:N%d" % last

    rng_n = "N%d:N%d" % (R0, last)
    for txt, bg, fg in [("工序延期", "FFC7CE", "9C0006"), ("进行中", "FFEB9C", "7F6000"),
                        ("未开工", "E7E6E6", "808080"), ("全部完成", "C6EFCE", "1F6E43"),
                        ("类目待确认", "FFF2CC", "C55A11")]:
        ws.conditional_formatting.add(rng_n, FormulaRule(
            formula=['EXACT($N%d,"%s")' % (R0, txt)], fill=fill(bg),
            font=Font(name=FONT_BASE, size=9, bold=True, color=fg), stopIfTrue=False))
    ws.conditional_formatting.add("L%d:L%d" % (R0, last),
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="63C384"))
    for i in range(1, 10):
        cd, ca = CL(col_due(i)), CL(col_act(i))
        ws.conditional_formatting.add("%s%d:%s%d" % (cd, R0, cd, last), FormulaRule(
            formula=['AND(ISNUMBER($%s%d),$%s%d<TODAY(),$%s%d="")' % (cd, R0, cd, R0, ca, R0)],
            fill=fill("FFC7CE"), font=Font(name=FONT_BASE, size=9, bold=True, color="9C0006"), stopIfTrue=False))
        ws.conditional_formatting.add("%s%d:%s%d" % (ca, R0, ca, last), FormulaRule(
            formula=['AND(ISNUMBER($%s%d),ISNUMBER($%s%d),$%s%d>$%s%d)' % (ca, R0, cd, R0, ca, R0, cd, R0)],
            fill=fill("FCE4D6"), font=Font(name=FONT_BASE, size=9, color="C55A11"), stopIfTrue=False))
        ws.conditional_formatting.add("%s%d:%s%d" % (ca, R0, ca, last), FormulaRule(
            formula=['AND(ISNUMBER($%s%d),ISNUMBER($%s%d),$%s%d<=$%s%d)' % (ca, R0, cd, R0, ca, R0, cd, R0)],
            fill=fill("C6EFCE"), font=Font(name=FONT_BASE, size=9, color="1F6E43"), stopIfTrue=False))
    return last
