# -*- coding: utf-8 -*-
"""分析视图：01驾驶舱 / 04甘特 / 05逾期专项 / 06合同汇总 / 00使用说明"""
import datetime
from openpyxl.utils import get_column_letter as CL
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, DataBarRule, ColorScaleRule
from build_part1 import *
from build_util import *

DATE_FMT = "yyyy-mm-dd"
R0 = 3
Q_ORD = q(SH_ORDER)
Q_PRC = q(SH_PROC)
Q_DICT = q(SH_DICT)
Q_CONT = q(SH_CONT)
Q_RAW = q(SH_RAW)
Q_PROD = q(SH_PROD)
Q_SHIP = q(SH_SHIP)
Q_SHIPCHK = q(SH_SHIPCHK)
Q_FCST = q(SH_FCST)
Q_REDUN = q(SH_REDUN)

# 02 订单明细 新列号（2026-08 改版：去掉母合同号/原始品类；新增合同总数；待交付=未到货总数量AG）
# 序号A 工厂B 合同号C SKU D 标准工艺类目E 订单总数量F 已交付数量G 待交付数量H 完成率I
# 合同总数J 数量档位K 是否新单L 跟单员M 合同日期N 计划出货日O 生产周期P 距出货Q
# 交期状态R 逾期天数S 风险等级T 工序进度U 当前在制工序V 卡点状态W 最新跟进日期X 跟进结论Y 跟进备注Z
COL = dict(序号="A", 工厂="B", 合同号="C", SKU="D", 标准工艺类目="E",
           订单总数量="F", 已交付数量="G", 待交付数量="H", 完成率="I", 合同总数="J",
           数量档位="K", 是否新单="L", 跟单员="M", 合同日期="N", 计划出货日="O",
           生产周期="P", 距出货="Q", 交期状态="R", 逾期天数="S", 风险等级="T",
           工序进度="U", 当前在制工序="V", 卡点状态="W", 最新跟进日期="X",
           跟进结论="Y", 跟进备注="Z")

def build_dash(wb, orders, last, dinfo):
    ws = wb.create_sheet(SH_DASH)
    def col(c):
        return "{o}!${c}${r0}:${c}${l}".format(o=Q_ORD, c=c, r0=R0, l=last)

    banner(ws, 1, 12, "跟单驾驶舱 · 全局体检     （所有数字实时联动 02订单明细，无需手动刷新）")
    ws.merge_cells("A2:L2")
    put(ws, 2, 1, '=" 数据基准日： "&TEXT(TODAY(),"yyyy年m月d日")&"    |    在制订单行 "&'
                  'COUNTIF({k},">0")&" 行 / 共 "&COUNTA({c})&" 行    |    口径：一行 = 一个「合同号 + SKU」"'
        .format(k=col(COL["待交付数量"]), c=col(COL["SKU"])),
        font=Font(name=FONT_BASE, size=10, bold=True, color="1F4E79"), fillc="EAF3FB", align=AL_L)
    ws.row_dimensions[2].height = 24

    # ---------- KPI ----------
    kpis = [
        ("订单总量(件)", "=SUM(%s)" % col(COL["订单总数量"]), "#,##0", "1F4E79"),
        ("已交付(件)", "=SUM(%s)" % col(COL["已交付数量"]), "#,##0", "1F6E43"),
        ("待交付(件)", "=SUM(%s)" % col(COL["待交付数量"]), "#,##0", "C55A11"),
        ("整体完成率", "=IFERROR(SUM({j})/SUM({h}),0)".format(
            j=col(COL["已交付数量"]), h=col(COL["订单总数量"])), "0.0%", "1F4E79"),
        ("逾期行数", '=COUNTIF({t},"已逾期")'.format(t=col(COL["交期状态"])), "#,##0", "9C0006"),
        ("逾期待交付(件)", '=SUMIF({t},"已逾期",{k})'.format(
            t=col(COL["交期状态"]), k=col(COL["待交付数量"])), "#,##0", "9C0006"),
        ("呆滞行数(>180天)", '=COUNTIF({v},"呆滞")'.format(v=col(COL["风险等级"])), "#,##0", "404040"),
        ("紧急行数(≤7天)", '=COUNTIF({t},"紧急")'.format(t=col(COL["交期状态"])), "#,##0", "C55A11"),
        ("在制母合同数", '=SUMPRODUCT(({c}!$H$5:$H$2000>0)*1)'.format(c=Q_CONT), "#,##0", "5B3A8E"),
        ("类目待确认行", '=COUNTIF({h},"待确认")'.format(h=col(COL["标准工艺类目"])), "#,##0", "C55A11"),
        ("在制订单行", '=COUNTIF({j},">0")'.format(j=col(COL["待交付数量"])), "#,##0", "808080"),
        ("平均生产周期(天)", "=ROUND(AVERAGE(%s),0)" % col(COL["生产周期"]), "0", "808080"),
    ]
    r = 4
    for i, (name, f, fmt, cc) in enumerate(kpis):
        c0 = 1 + (i % 6) * 2
        rr = r + (i // 6) * 3
        ws.merge_cells(start_row=rr, start_column=c0, end_row=rr, end_column=c0+1)
        put(ws, rr, c0, name, font=Font(name=FONT_BASE, size=9, color="FFFFFF"), fillc=cc, align=AL_C)
        ws.cell(row=rr, column=c0+1).fill = fill(cc)
        ws.merge_cells(start_row=rr+1, start_column=c0, end_row=rr+1, end_column=c0+1)
        put(ws, rr+1, c0, f, font=Font(name=FONT_BASE, size=16, bold=True, color=cc),
            fillc=F_KPI, align=AL_C, numfmt=fmt)
        ws.row_dimensions[rr].height = 18
        ws.row_dimensions[rr+1].height = 30

    r = 11

    def block(title, label, keys, key_col, extra=None, note=None):
        nonlocal r
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        put(ws, r, 1, title, font=f_title(11), fillc=F_KPI, align=AL_L)
        r += 1
        hd = [label, "SKU行数", "订单总量", "已交付", "待交付", "完成率", "逾期行", "呆滞行", "最久逾期(天)"]
        for j, t in enumerate(hd, start=1):
            put(ws, r, j, t, font=f_head(9), fillc=F_SUBHEAD)
        r += 1
        start = r
        for k in keys:
            put(ws, r, 1, k, font=f_input(9))
            put(ws, r, 2, '=COUNTIF({c},$A{r})'.format(c=col(key_col), r=r), font=f_formula(9), numfmt="#,##0")
            put(ws, r, 3, '=SUMIF({c},$A{r},{v})'.format(c=col(key_col), r=r, v=col(COL["订单总数量"])), font=f_formula(9), numfmt="#,##0")
            put(ws, r, 4, '=SUMIF({c},$A{r},{v})'.format(c=col(key_col), r=r, v=col(COL["已交付数量"])), font=f_formula(9), numfmt="#,##0")
            put(ws, r, 5, '=SUMIF({c},$A{r},{v})'.format(c=col(key_col), r=r, v=col(COL["待交付数量"])), font=f_formula(9), numfmt="#,##0")
            put(ws, r, 6, '=IFERROR($D{r}/$C{r},0)'.format(r=r), font=f_formula(9), numfmt="0.0%")
            put(ws, r, 7, '=COUNTIFS({c},$A{r},{t},"已逾期")'.format(c=col(key_col), r=r, t=col(COL["交期状态"])), font=f_formula(9), numfmt="#,##0")
            put(ws, r, 8, '=COUNTIFS({c},$A{r},{v},"呆滞")'.format(c=col(key_col), r=r, v=col(COL["风险等级"])), font=f_formula(9), numfmt="#,##0")
            put(ws, r, 9, '=SUMPRODUCT(MAX(({c}=$A{r})*({u})))'.format(u=col(COL["逾期天数"]), c=col(key_col), r=r), font=f_formula(9), numfmt="#,##0")
            r += 1
        # 合计
        put(ws, r, 1, "合计", font=f_formula(9, True), fillc=F_GREY)
        for j, ltr in zip(range(2, 10), "BCDEFGHI"):
            if j == 6:
                put(ws, r, j, '=IFERROR($D{r}/$C{r},0)'.format(r=r), font=f_formula(9, True), fillc=F_GREY, numfmt="0.0%")
            elif j == 9:
                put(ws, r, j, "=MAX({c}{s}:{c}{e})".format(c=ltr, s=start, e=r-1), font=f_formula(9, True), fillc=F_GREY, numfmt="#,##0")
            else:
                put(ws, r, j, "=SUM({c}{s}:{c}{e})".format(c=ltr, s=start, e=r-1), font=f_formula(9, True), fillc=F_GREY, numfmt="#,##0")
        ws.conditional_formatting.add("F%d:F%d" % (start, r-1),
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="63C384"))
        r += 1
        if note:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
            put(ws, r, 1, note, font=f_note(9), align=AL_L, border=False)
            r += 1

    facs = dinfo["lists"]["工厂"]
    block("① 按工厂", "工厂", facs, COL["工厂"],
          note="※ 完成率低 + 逾期行多 = 优先约谈对象；最久逾期天数 > 180 天的工厂需启动清尾评审。")
    block("② 按标准工艺类目", "标准工艺类目", CAT_ORDER + ["待确认"], COL["标准工艺类目"],
          note="※「待确认」= 产品资料表里该产品编号无工艺品类（留空或不在表），请在「产品资料表」补全后即可自动排期。")
    block("③ 按跟单员", "跟单员", dinfo["lists"]["跟单员"], COL["跟单员"])

    # 状态分布
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    put(ws, r, 1, "④ 状态分布（交期 / 风险 / 生产卡点）", font=f_title(11), fillc=F_KPI, align=AL_L)
    r += 1
    hd = ["交期状态", "行数", "待交付件数", "", "风险等级", "行数", "", "生产卡点", "行数"]
    for j, t in enumerate(hd, start=1):
        put(ws, r, j, t, font=f_head(9), fillc=F_SUBHEAD if t else "FFFFFF", border=bool(t))
    r += 1
    st1 = ["已逾期", "紧急", "预警", "正常", "已交付"]
    st2 = ["呆滞", "高", "中", "低", "无"]
    st3 = ["工序延期", "进行中", "未开工", "全部完成", "类目待确认"]
    for i in range(5):
        put(ws, r+i, 1, st1[i], font=f_input(9))
        put(ws, r+i, 2, '=COUNTIF({t},$A{r})'.format(t=col(COL["交期状态"]), r=r+i), font=f_formula(9), numfmt="#,##0")
        put(ws, r+i, 3, '=SUMIF({t},$A{r},{k})'.format(t=col(COL["交期状态"]), r=r+i, k=col(COL["待交付数量"])), font=f_formula(9), numfmt="#,##0")
        ws.cell(row=r+i, column=4).border = NOBORDER
        put(ws, r+i, 5, st2[i], font=f_input(9))
        put(ws, r+i, 6, '=COUNTIF({v},$E{r})'.format(v=col(COL["风险等级"]), r=r+i), font=f_formula(9), numfmt="#,##0")
        ws.cell(row=r+i, column=7).border = NOBORDER
        put(ws, r+i, 8, st3[i], font=f_input(9))
        put(ws, r+i, 9, '=COUNTIF({y},$H{r})'.format(y=col(COL["卡点状态"]), r=r+i), font=f_formula(9), numfmt="#,##0")
    r += 6

    # 逾期账龄
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    put(ws, r, 1, "⑤ 逾期账龄分布（只统计「待交付 > 0 且已过计划出货日」的行）", font=f_title(11), fillc=F_KPI, align=AL_L)
    r += 1
    for j, t in enumerate(["逾期区间", "行数", "待交付件数", "占逾期件数比", "处理建议"], start=1):
        put(ws, r, j, t, font=f_head(9), fillc=F_SUBHEAD)
    r += 1
    ages = [("1-30天", 1, 30, "正常催办，工厂给出补产日期"),
            ("31-90天", 31, 90, "升级到工厂负责人，书面确认新交期"),
            ("91-180天", 91, 180, "评估是否改单/转产/让步接收"),
            (">180天(呆滞)", 181, 99999, "启动清尾：核销、退单或转内销，别再挂账")]
    ast = r
    for nm, lo, hi, adv in ages:
        put(ws, r, 1, nm, font=f_input(9))
        put(ws, r, 2, '=COUNTIFS({u},">="&{lo},{u},"<="&{hi},{k},">0")'.format(
            u=col(COL["逾期天数"]), lo=lo, hi=hi, k=col(COL["待交付数量"])),
            font=f_formula(9), numfmt="#,##0")
        put(ws, r, 3, '=SUMIFS({k},{u},">="&{lo},{u},"<="&{hi},{k},">0")'.format(
            k=col(COL["待交付数量"]), u=col(COL["逾期天数"]), lo=lo, hi=hi),
            font=f_formula(9), numfmt="#,##0")
        put(ws, r, 4, '=IFERROR($C{r}/SUM($C${s}:$C${e}),0)'.format(r=r, s=ast, e=ast+3), font=f_formula(9), numfmt="0.0%")
        put(ws, r, 5, adv, font=f_note(9, "808080", False), align=AL_L)
        r += 1
    ws.conditional_formatting.add("D%d:D%d" % (ast, ast+3),
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="F8696B"))

    # ⑥ 发货需求满足度（窗口内，详情见「10发货需求判断」）
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    put(ws, r, 1, "⑥ 发货需求满足度（窗口内 · 详见「10发货需求判断」）", font=f_title(11), fillc=F_KPI, align=AL_L)
    r += 1
    sc = [("窗口内需求总数量", "=SUM({c}!E3)".format(c=Q_SHIPCHK), "#,##0"),
          ("可供应未入库总数量", "=SUM({c}!E4)".format(c=Q_SHIPCHK), "#,##0"),
          ("总缺口(供应-需求)", "=SUM({c}!E5)".format(c=Q_SHIPCHK), "#,##0"),
          ("缺口行数", "=SUM({c}!E6)".format(c=Q_SHIPCHK), "#,##0"),
          ("满足率(有需求行)", "=SUM({c}!E7)".format(c=Q_SHIPCHK), "0.0%")]
    for i, (nm, f, fmt) in enumerate(sc):
        rr = r + i
        put(ws, rr, 1, nm, font=f_input(9))
        put(ws, rr, 2, f, font=f_formula(11, True), fillc=F_KPI, numfmt=fmt)
        put(ws, rr, 3, "", border=False)
    put(ws, r, 4, "供应口径 = 原始数据表「未入库数量」(按产品编号+团队)；需求 = 发货需求表窗口内「发货数量」(按产品编号+团队)。",
        font=f_note(9, "808080", False), align=AL_L)
    ws.merge_cells(start_row=r, start_column=4, end_row=r+4, end_column=9)
    r += 5

    # ⑦ 合同冗余（详见「11合同冗余」）
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    put(ws, r, 1, "⑦ 合同冗余（超期未交付占压 · 详见「11合同冗余」）", font=f_title(11), fillc=F_KPI, align=AL_L)
    r += 1
    sc7 = [("在手待交付总量(件)", "=SUM({c}!E3)".format(c=Q_REDUN), "#,##0"),
           ("在手待交付金额(元)", "=SUM({c}!E4)".format(c=Q_REDUN), "#,##0"),
           ("冗余数量(超阈值,件)", "=SUM({c}!E5)".format(c=Q_REDUN), "#,##0"),
           ("冗余金额(元)", "=SUM({c}!E6)".format(c=Q_REDUN), "#,##0"),
           ("冗余金额占比", "=SUM({c}!E7)".format(c=Q_REDUN), "0.0%"),
           ("冗余行数", "=SUM({c}!E8)".format(c=Q_REDUN), "#,##0"),
           ("不可接单组合数", "=SUM({c}!E9)".format(c=Q_REDUN), "#,##0")]
    for i, (nm, f, fmt) in enumerate(sc7):
        rr = r + i
        put(ws, rr, 1, nm, font=f_input(9))
        put(ws, rr, 2, f, font=f_formula(11, True), fillc=F_KPI, numfmt=fmt)
        put(ws, rr, 3, "", border=False)
    put(ws, r, 4, "口径：交付周期 = 基准日 − 交货日期；冗余 = 未到货总数量(超阈值) × 采购单价。"
                   "⑤接单判断需先在「销售预测表」填未来3月预计出货。",
        font=f_note(9, "808080", False), align=AL_L)
    ws.merge_cells(start_row=r, start_column=4, end_row=r+6, end_column=9)
    r += 7

    setw(ws, [18, 12, 12, 12, 12, 12, 12, 14, 14, 12, 12, 12])
    ws.column_dimensions["E"].width = 14
    ws.sheet_view.showGridLines = False


# ================= 04 甘特视图 =================
NWEEK = 22
def build_gantt(wb, last):
    ws = wb.create_sheet(SH_GANTT)
    banner(ws, 1, 8 + NWEEK, "工序甘特视图 · 滚动 22 周（自动以本周为基准：前 4 周 → 后 17 周）")
    ws.merge_cells("A2:H2")
    put(ws, 2, 1, "格子里的数字 = 到该周末为止「计划应完成开工」的工序道数（1~9），颜色由浅到深即进度推进；"
                  "红底 = 该周已超出计划出货日仍未交清。用左侧筛选按工厂/合同筛看。",
        font=f_note(9, "1F4E79", False), fillc="EAF3FB", align=AL_L)
    ws.row_dimensions[2].height = 30

    fixed = [("序号", 6), ("工厂", 9), ("合同号", 24), ("SKU", 20),
             ("类目", 13), ("待交付", 9), ("计划出货日", 12), ("卡点状态", 11)]
    for j, (nm, w) in enumerate(fixed, start=1):
        put(ws, 3, j, nm, font=f_head(9), fillc=F_HEADER)
        ws.column_dimensions[CL(j)].width = w
    # 周表头
    for i in range(NWEEK):
        c = 9 + i
        f = "=TODAY()-WEEKDAY(TODAY(),3)-28" if i == 0 else "=%s3+7" % CL(c-1)
        put(ws, 3, c, f, font=f_head(8), fillc="2E75B6", numfmt="m/d")
        ws.column_dimensions[CL(c)].width = 5.6
    ws.row_dimensions[3].height = 30

    P_PLAN = "{p}!${a}{{r}}:${b}{{r}}".format(p=Q_PRC, a="AP", b="AX")
    for r in range(4, 4 + (last - R0 + 1)):
        sr = R0 + (r - 4)
        put(ws, r, 1, "={o}!A{s}".format(o=Q_ORD, s=sr), font=f_xref(8))
        put(ws, r, 2, "={o}!B{s}".format(o=Q_ORD, s=sr), font=f_xref(8))
        put(ws, r, 3, "={o}!C{s}".format(o=Q_ORD, s=sr), font=f_xref(8), align=AL_L)
        put(ws, r, 4, "={o}!D{s}".format(o=Q_ORD, s=sr), font=f_xref(8), align=AL_L)      # SKU
        put(ws, r, 5, "={o}!E{s}".format(o=Q_ORD, s=sr), font=f_xref(8))                   # 标准工艺类目
        put(ws, r, 6, "={o}!H{s}".format(o=Q_ORD, s=sr), font=f_xref(8, True), numfmt="#,##0")  # 待交付
        put(ws, r, 7, "={o}!O{s}".format(o=Q_ORD, s=sr), font=f_xref(8), numfmt=DATE_FMT)  # 计划出货日
        put(ws, r, 8, "={o}!W{s}".format(o=Q_ORD, s=sr), font=f_xref(8, True))             # 卡点状态
        rng = P_PLAN.format(r=sr)
        for i in range(NWEEK):
            c = 9 + i
            f = ('=IF(COUNT({rg})=0,"",IF(SUMPRODUCT(({rg}<={cl}$3+6)*ISNUMBER({rg}))=0,"",'
                 'SUMPRODUCT(({rg}<={cl}$3+6)*ISNUMBER({rg}))))').format(rg=rng, cl=CL(c))
            put(ws, r, c, f, font=Font(name=FONT_BASE, size=8, color="404040"))
        ws.row_dimensions[r].height = 14

    lastg = 3 + (last - R0 + 1)
    ws.freeze_panes = "I4"
    ws.auto_filter.ref = "A3:H%d" % lastg
    g = "I4:%s%d" % (CL(8+NWEEK), lastg)
    ws.conditional_formatting.add(g, ColorScaleRule(
        start_type="num", start_value=1, start_color="DEEBF7",
        mid_type="num", mid_value=5, mid_color="9DC3E6",
        end_type="num", end_value=9, end_color="1F4E79"))
    ws.conditional_formatting.add(g, FormulaRule(
        formula=['AND(ISNUMBER($G4),I$3>$G4,$F4>0)'], fill=fill("FFC7CE"), stopIfTrue=False))
    ws.conditional_formatting.add("I3:%s3" % CL(8+NWEEK), FormulaRule(
        formula=['AND(I$3<=TODAY(),I$3+6>=TODAY())'],
        fill=fill("FFC000"), font=Font(name=FONT_BASE, size=8, bold=True, color="000000"), stopIfTrue=False))
    ws.sheet_view.showGridLines = False


# ================= 05 逾期呆滞专项 =================
def build_late(wb, orders, dinfo):
    ws = wb.create_sheet(SH_LATE)
    TODAY = datetime.date.today()
    idx = [(i, o) for i, o in enumerate(orders) if o["rem"] > 0 and o["sdate"] and o["sdate"] < TODAY]
    idx.sort(key=lambda t: -(TODAY - t[1]["sdate"]).days)

    banner(ws, 1, 16, "逾期 & 呆滞订单专项催办清单（生成时按逾期天数降序，共 %d 行）" % len(idx))
    ws.merge_cells("A2:P2")
    put(ws, 2, 1, "本表所有灰色字段都是从 02订单明细 实时取数，改主表这里自动更新；橙色底的 4 列请你手工维护。"
                  "「仍逾期?」变成「已解决」的行可以隐藏。排序顺序在生成时固定，需要重排请点表头筛选按钮。",
        font=f_note(9, "1F4E79", False), fillc="EAF3FB", align=AL_L)
    ws.row_dimensions[2].height = 30

    hd = [("#", 5), ("仍逾期?", 9), ("逾期天数", 9), ("账龄", 12), ("风险", 8), ("工厂", 9),
          ("合同号", 24), ("SKU", 20), ("订单量", 9), ("待交付", 9), ("完成率", 9),
          ("计划出货日", 12), ("跟单员", 9), ("当前在制工序", 14),
          ("★催办动作", 26), ("★工厂承诺交期", 13), ("★跟进结论", 16), ("★备注", 30)]
    for j, (nm, w) in enumerate(hd, start=1):
        put(ws, 3, j, nm, font=f_head(9), fillc="9C0006" if nm.startswith("★") else F_HEADER)
        ws.column_dimensions[CL(j)].width = w
    ws.row_dimensions[3].height = 30

    r = 4
    for rank, (i, o) in enumerate(idx, start=1):
        sr = R0 + i
        put(ws, r, 1, rank, font=f_formula(8))
        put(ws, r, 2, '=IF({o}!H{s}<=0,"已解决",IF({o}!Q{s}<0,"仍逾期","已改期"))'.format(o=Q_ORD, s=sr), font=f_xref(8, True))
        put(ws, r, 3, "={o}!S{s}".format(o=Q_ORD, s=sr), font=f_xref(8, True), numfmt="#,##0")
        put(ws, r, 4, '=IF($C{r}=0,"—",IF($C{r}<=30,"1-30天",IF($C{r}<=90,"31-90天",IF($C{r}<=180,"91-180天",">180天呆滞"))))'.format(r=r), font=f_formula(8))
        put(ws, r, 5, "={o}!T{s}".format(o=Q_ORD, s=sr), font=f_xref(8, True))
        put(ws, r, 6, "={o}!B{s}".format(o=Q_ORD, s=sr), font=f_xref(8))
        put(ws, r, 7, "={o}!C{s}".format(o=Q_ORD, s=sr), font=f_xref(8), align=AL_L)
        put(ws, r, 8, "={o}!D{s}".format(o=Q_ORD, s=sr), font=f_xref(8), align=AL_L)
        put(ws, r, 9, "={o}!F{s}".format(o=Q_ORD, s=sr), font=f_xref(8), numfmt="#,##0")
        put(ws, r, 10, "={o}!H{s}".format(o=Q_ORD, s=sr), font=f_xref(8, True), numfmt="#,##0")
        put(ws, r, 11, "={o}!I{s}".format(o=Q_ORD, s=sr), font=f_xref(8), numfmt="0.0%")
        put(ws, r, 12, "={o}!O{s}".format(o=Q_ORD, s=sr), font=f_xref(8), numfmt=DATE_FMT)
        put(ws, r, 13, "={o}!M{s}".format(o=Q_ORD, s=sr), font=f_xref(8))
        put(ws, r, 14, "={o}!V{s}".format(o=Q_ORD, s=sr), font=f_xref(8))
        for c in (15, 16, 17, 18):
            put(ws, r, c, None, font=f_input(9), fillc="FFF0E1",
                numfmt=DATE_FMT if c == 16 else None, align=AL_L if c in (15, 18) else AL_C)
        r += 1
    lastr = r - 1

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = "A3:R%d" % lastr
    ls = dinfo["list_start"]
    cnt = len(dinfo["lists"]["跟进结论"])
    dv = DataValidation(type="list", formula1="={d}!$F${s}:$F${e}".format(d=Q_DICT, s=ls, e=ls+cnt-1),
                        allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv)
    dv.add("Q4:Q%d" % lastr)

    for txt, bg, fg in [("仍逾期", "FFC7CE", "9C0006"), ("已解决", "C6EFCE", "1F6E43"), ("已改期", "FFEB9C", "7F6000")]:
        ws.conditional_formatting.add("B4:B%d" % lastr, FormulaRule(
            formula=['EXACT($B4,"%s")' % txt], fill=fill(bg),
            font=Font(name=FONT_BASE, size=8, bold=True, color=fg), stopIfTrue=False))
    ws.conditional_formatting.add("A4:R%d" % lastr, FormulaRule(
        formula=['$D4=">180天呆滞"'], fill=fill("F2DCDB"), stopIfTrue=False))
    ws.conditional_formatting.add("C4:C%d" % lastr,
        DataBarRule(start_type="num", start_value=0, end_type="max", color="F8696B"))
    return lastr


# ================= 06 合同级汇总 =================
def build_contract(wb, orders, last):
    ws = wb.create_sheet(SH_CONT)
    seen, mains = set(), []
    for o in orders:
        m = o["con"]   # 母合同号 = 合同号全文（HB 非补单，不再剥离 -HB）
        if m not in seen:
            seen.add(m); mains.append((m, o["fac"]))

    banner(ws, 1, 16, "合同级汇总（按「母合同号」归集，母合同号=合同号全文；共 %d 个合同）" % len(mains))
    ws.merge_cells("A2:P2")
    put(ws, 2, 1, "回答一个问题：这张合同到底能不能整单出？「缺口件数」为 0 才可以约柜。"
                  "所有数字由 02订单明细 汇总而来，主表一改这里自动跟着变。",
        font=f_note(9, "1F4E79", False), fillc="EAF3FB", align=AL_L)
    ws.row_dimensions[2].height = 26

    hd = [("#", 5), ("母合同号", 22), ("工厂", 9), ("含单据行", 9), ("含SKU种类", 10),
          ("订单总量", 10), ("已交付", 10), ("待交付", 10), ("完成率", 10),
          ("最早合同日", 12), ("最紧出货日", 12), ("距最紧交期(天)", 12),
          ("合同交期状态", 12), ("最大逾期天数", 12), ("整单可出货?", 16), ("含呆滞行", 10)]
    for j, (nm, w) in enumerate(hd, start=1):
        put(ws, j and 4 or 4, j, nm, font=f_head(9), fillc=F_HEADER)
        ws.column_dimensions[CL(j)].width = w
    ws.row_dimensions[4].height = 30
    put(ws, 3, 1, "", border=False)

    def c(x):
        return "{o}!${x}${r}:${x}${l}".format(o=Q_ORD, x=x, r=R0, l=last)
    D = c("C")   # 合同号(=母合同号全文)
    CC = c("C")  # 合同号
    E = c("D")   # SKU
    H = c("F")   # 订单总数量
    I = c("G")   # 已交付数量
    J = c("H")   # 待交付数量
    O = c("N")   # 合同日期
    P = c("O")   # 计划出货日
    T = c("S")   # 逾期天数
    U = c("T")   # 风险等级
    r = 5
    for i, (m, fac) in enumerate(mains, start=1):
        put(ws, r, 1, i, font=f_formula(8))
        put(ws, r, 2, m, font=f_input(9), align=AL_L)
        put(ws, r, 3, fac, font=f_input(9))
        put(ws, r, 4, '=COUNTIF({d},$B{r})'.format(d=D, r=r), font=f_formula(8), numfmt="0")
        put(ws, r, 5, '=SUMPRODUCT(({d}=$B{r})/COUNTIFS({d},{d},{e},{e}))'.format(d=D, r=r, e=E),
            font=f_formula(8), numfmt="0")
        put(ws, r, 6, "=SUMIF({d},$B{r},{v})".format(d=D, r=r, v=H), font=f_formula(9), numfmt="#,##0")
        put(ws, r, 7, "=SUMIF({d},$B{r},{v})".format(d=D, r=r, v=I), font=f_formula(9), numfmt="#,##0")
        put(ws, r, 8, "=SUMIF({d},$B{r},{v})".format(d=D, r=r, v=J), font=f_formula(9, True), numfmt="#,##0")
        put(ws, r, 9, "=IFERROR($G{r}/$F{r},0)".format(r=r), font=f_formula(9), numfmt="0.0%")
        put(ws, r, 10, "=SUMPRODUCT(MIN(({d}=$B{r})*{v}+({d}<>$B{r})*99999))".format(v=O, d=D, r=r),
            font=f_formula(8), numfmt=DATE_FMT)
        put(ws, r, 11, "=SUMPRODUCT(MIN(({d}=$B{r})*{v}+({d}<>$B{r})*99999))".format(v=P, d=D, r=r),
            font=f_formula(8), numfmt=DATE_FMT)
        put(ws, r, 12, "=$L{r}-TODAY()".format(r=r), font=f_formula(8), numfmt="0")
        put(ws, r, 13, '=IF($H{r}<=0,"已交清",IF($L{r}<0,"已逾期",IF($L{r}<=7,"紧急",IF($L{r}<=15,"预警","正常"))))'.format(r=r),
            font=f_formula(9, True))
        put(ws, r, 14, "=SUMPRODUCT(MAX(({d}=$B{r})*{v}))".format(v=T, d=D, r=r), font=f_formula(8), numfmt="0")
        put(ws, r, 15, '=IF($H{r}<=0,"✔ 可整单出货","缺 "&TEXT($H{r},"#,##0")&" 件 / "&COUNTIFS({d},$B{r},{k},">0")&" 个SKU")'
            .format(r=r, d=D, k=J), font=f_formula(9, True), align=AL_L)
        put(ws, r, 16, '=COUNTIFS({d},$B{r},{v},"呆滞")'.format(d=D, r=r, v=U), font=f_formula(8), numfmt="0")
        r += 1
    lastc = r - 1
    put(ws, r, 1, "合计", font=f_formula(9, True), fillc=F_GREY)
    put(ws, r, 2, "", fillc=F_GREY)
    put(ws, r, 3, "", fillc=F_GREY)
    for j, ltr in [(4, "D"), (5, "E"), (6, "F"), (7, "G"), (8, "H"), (16, "P")]:
        put(ws, r, j, "=SUM({c}5:{c}{e})".format(c=ltr, e=lastc), font=f_formula(9, True), fillc=F_GREY, numfmt="#,##0")
    put(ws, r, 9, "=IFERROR($G{r}/$F{r},0)".format(r=r), font=f_formula(9, True), fillc=F_GREY, numfmt="0.0%")
    for j in (10, 11, 12, 13, 14, 15):
        put(ws, r, j, "", fillc=F_GREY)

    ws.freeze_panes = "C5"
    ws.auto_filter.ref = "A4:P%d" % lastc
    for txt, bg, fg in [("已逾期", "FFC7CE", "9C0006"), ("紧急", "FFD9B3", "974706"),
                        ("预警", "FFEB9C", "7F6000"), ("正常", "C6EFCE", "1F6E43"),
                        ("已交清", "E7E6E6", "808080")]:
        ws.conditional_formatting.add("M5:M%d" % lastc, FormulaRule(
            formula=['EXACT($M5,"%s")' % txt], fill=fill(bg),
            font=Font(name=FONT_BASE, size=9, bold=True, color=fg), stopIfTrue=False))
    ws.conditional_formatting.add("I5:I%d" % lastc,
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="63C384"))
    ws.conditional_formatting.add("O5:O%d" % lastc, FormulaRule(
        formula=['$H5<=0'], fill=fill("C6EFCE"),
        font=Font(name=FONT_BASE, size=9, bold=True, color="1F6E43"), stopIfTrue=False))
    return lastc


# ================= 10 发货需求判断 =================
def build_shipchk(wb, ship_count):
    ws = wb.create_sheet(SH_SHIPCHK)
    NB = 600                      # 显示行缓冲（发货需求表在此范围内增删行会自动反映）
    DSTART, DEND = 9, 9 + NB - 1  # 明细起始/结束行
    # 原始数据表 范围（产品编号C / 团队D / 合同数量AD / 未入库数量AQ）
    RSKU = "%s!$C$3:$C$804" % Q_RAW
    RTEAM = "%s!$D$3:$D$804" % Q_RAW
    RQTY = "%s!$AD$3:$AD$804" % Q_RAW
    RUNSTK = "%s!$AQ$3:$AQ$804" % Q_RAW
    # 发货需求表 范围（SKU A / 团队 B / 发货数量 F / 发货时间 G）
    SSKU = "%s!$A$4:$A$804" % Q_SHIP
    STEAM = "%s!$B$4:$B$804" % Q_SHIP
    SQTY = "%s!$F$4:$F$804" % Q_SHIP
    STIME = "%s!$G$4:$G$804" % Q_SHIP

    banner(ws, 1, 14, "发货需求判断 · 我的未入库数量能否满足发货需求？")
    ws.merge_cells("A2:N2")
    put(ws, 2, 1, "逻辑：供应 = 原始数据表「未入库数量」按 产品编号+团队 汇总；需求 = 发货需求表「发货数量」按 产品编号+团队 在发货窗口内汇总。"
                  "差额≥0 即满足。设好下方起止日即可看特定时段缺口。",
        font=f_note(9, "1F4E79", False), fillc="EAF3FB", align=AL_L)
    ws.row_dimensions[2].height = 30

    # ---- 参数块 ----
    put(ws, 3, 1, "发货开始日(留空=不限制)", font=f_note(9, "808080", True), align=AL_L)
    put(ws, 3, 2, None, font=f_input(9), fillc=F_ASSUME, numfmt=DATE_FMT)
    put(ws, 4, 1, "发货结束日(留空=不限制)", font=f_note(9, "808080", True), align=AL_L)
    put(ws, 4, 2, None, font=f_input(9), fillc=F_ASSUME, numfmt=DATE_FMT)
    put(ws, 5, 1, "有效开始日(自动)", font=f_note(9, "808080", False), align=AL_L)
    put(ws, 5, 2, '=IF($B$3="",DATE(2000,1,1),$B$3)', font=f_formula(9), numfmt=DATE_FMT)
    put(ws, 6, 1, "有效结束日(自动)", font=f_note(9, "808080", False), align=AL_L)
    put(ws, 6, 2, '=IF($B$4="",DATE(2999,12,31),$B$4)', font=f_formula(9), numfmt=DATE_FMT)

    # ---- KPI 汇总（被驾驶舱⑥引用：E3..E7）----
    kpis = [("窗口内需求总数量", "=SUM(I%d:I%d)" % (DSTART, DEND), "#,##0"),
            ("可供应未入库总数量", "=SUM(J%d:J%d)" % (DSTART, DEND), "#,##0"),
            ("总缺口(供应-需求)", "=E4-E3", "#,##0"),
            ("缺口行数", '=COUNTIF(L%d:L%d,"<0")' % (DSTART, DEND), "#,##0"),
            ("满足率(有需求行)", '=IFERROR(COUNTIFS(I%d:I%d,">0",L%d:L%d,">=0")/COUNTIF(I%d:I%d,">0"),0)' % (DSTART, DEND, DSTART, DEND, DSTART, DEND), "0.0%")]
    for i, (nm, f, fmt) in enumerate(kpis):
        put(ws, 3 + i, 4, nm, font=Font(name=FONT_BASE, size=9, color="FFFFFF"), fillc="1F4E79", align=AL_L)
        put(ws, 3 + i, 5, f, font=Font(name=FONT_BASE, size=14, bold=True, color="1F4E79"), fillc=F_KPI, align=AL_C, numfmt=fmt)

    # ---- 明细表头 ----
    hd = [("序号", 6), ("产品编号", 22), ("团队", 18), ("运营", 12), ("工厂", 14), ("跟单", 12),
          ("发货时间", 12), ("发货数量", 11), ("窗口内需求数量", 14), ("可供应未入库", 13),
          ("合同数量(汇总)", 13), ("差额", 11), ("满足状态", 14), ("主行", 9)]
    for j, (nm, w) in enumerate(hd, start=1):
        put(ws, 8, j, nm, font=f_head(9), fillc=F_HEADER)
        ws.column_dimensions[CL(j)].width = w
    ws.row_dimensions[8].height = 30

    for k in range(NB):
        r = DSTART + k
        sr = 4 + k                     # 对应 发货需求表 数据行
        # 直接引用发货需求表
        put(ws, r, 1, k + 1, font=f_formula(8))
        put(ws, r, 2, "=%s!A%d" % (Q_SHIP, sr), font=f_xref(9), align=AL_L)
        put(ws, r, 3, "=%s!B%d" % (Q_SHIP, sr), font=f_xref(9), align=AL_L)
        put(ws, r, 4, "=%s!C%d" % (Q_SHIP, sr), font=f_xref(9), align=AL_L)
        put(ws, r, 5, "=%s!D%d" % (Q_SHIP, sr), font=f_xref(9), align=AL_L)
        put(ws, r, 6, "=%s!E%d" % (Q_SHIP, sr), font=f_xref(9), align=AL_L)
        put(ws, r, 7, "=%s!G%d" % (Q_SHIP, sr), font=f_xref(9), numfmt=DATE_FMT)
        put(ws, r, 8, "=%s!F%d" % (Q_SHIP, sr), font=f_xref(9), numfmt="#,##0")
        # 汇总（按 产品编号+团队）
        put(ws, r, 9, ('=IF($B%d="","",SUMIFS(%s,%s,$B%d,%s,$C%d,%s,">="&$B$5,%s,"<="&$B$6))'
                       % (r, SQTY, SSKU, r, STEAM, r, STIME, STIME)),
            font=f_formula(9), numfmt="#,##0")
        put(ws, r, 10, '=IF($B%d="","",SUMIFS(%s,%s,$B%d,%s,$C%d))'
            % (r, RUNSTK, RSKU, r, RTEAM, r), font=f_formula(9), numfmt="#,##0")
        put(ws, r, 11, '=IF($B%d="","",SUMIFS(%s,%s,$B%d,%s,$C%d))'
            % (r, RQTY, RSKU, r, RTEAM, r), font=f_formula(9, False), numfmt="#,##0")
        put(ws, r, 12, '=IF($B%d="","",J%d-I%d)' % (r, r, r), font=f_formula(9, True), numfmt="#,##0")
        put(ws, r, 13, ('=IF($B%d="","",IF(I%d=0,"无需求",IF(L%d>=0,"满足✔",'
                         '"缺口 "&TEXT(L%d,"#,##0"))))') % (r, r, r, r),
            font=f_formula(9, True))
        put(ws, r, 14, '=IF($B%d="","",IF(COUNTIF($B$%d:$B%d,$B%d&"|"&$C%d)=1,"★主行",""))'
            % (r, DSTART, r, r, r), font=f_formula(8))

    # ---- 条件格式 ----
    ws.conditional_formatting.add("L%d:L%d" % (DSTART, DEND), FormulaRule(
        formula=['$L%d<0' % DSTART], fill=fill("FFC7CE"), font=Font(name=FONT_BASE, size=9, bold=True, color="9C0006"), stopIfTrue=False))
    ws.conditional_formatting.add("M%d:M%d" % (DSTART, DEND), FormulaRule(
        formula=['ISNUMBER(SEARCH("缺口",$M%d))' % DSTART], fill=fill("FFC7CE"),
        font=Font(name=FONT_BASE, size=9, bold=True, color="9C0006"), stopIfTrue=False))
    ws.conditional_formatting.add("M%d:M%d" % (DSTART, DEND), FormulaRule(
        formula=['EXACT($M%d,"满足✔")' % DSTART], fill=fill("C6EFCE"),
        font=Font(name=FONT_BASE, size=9, color="1F6E43"), stopIfTrue=False))

    setw(ws, [20])
    ws.freeze_panes = "C9"
    ws.auto_filter.ref = "A8:N%d" % DEND
    ws.sheet_view.showGridLines = False
    return DEND
