# -*- coding: utf-8 -*-
"""样式工具"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from build_part1 import (C_INPUT, C_FORMULA, C_XREF, C_WARN, F_HEADER, F_SUBHEAD,
                         F_ASSUME, F_INPUT, F_GREY, F_KPI)

THIN = Side(style="thin", color="BFBFBF")
MED = Side(style="medium", color="1F4E79")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NOBORDER = Border()

FONT_BASE = "微软雅黑"

def f_title(sz=14, color="1F4E79"):
    return Font(name=FONT_BASE, size=sz, bold=True, color=color)

def f_head(sz=10):
    return Font(name=FONT_BASE, size=sz, bold=True, color="FFFFFF")

def f_input(sz=10, bold=False):
    return Font(name=FONT_BASE, size=sz, color=C_INPUT, bold=bold)

def f_formula(sz=10, bold=False):
    return Font(name=FONT_BASE, size=sz, color=C_FORMULA, bold=bold)

def f_xref(sz=10, bold=False):
    return Font(name=FONT_BASE, size=sz, color=C_XREF, bold=bold)

def f_note(sz=9, color="808080", italic=True):
    return Font(name=FONT_BASE, size=sz, color=color, italic=italic)

def f_warn(sz=10, bold=True):
    return Font(name=FONT_BASE, size=sz, color=C_WARN, bold=bold)

def fill(hexc):
    return PatternFill("solid", fgColor=hexc)

AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
AL_LT = Alignment(horizontal="left", vertical="top", wrap_text=True)
AL_R = Alignment(horizontal="right", vertical="center")

def put(ws, row, col, value, font=None, fillc=None, align=None, numfmt=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or Font(name=FONT_BASE, size=10)
    if fillc:
        c.fill = fill(fillc)
    c.alignment = align or AL_C
    if numfmt:
        c.number_format = numfmt
    if border:
        c.border = BOX
    return c

def paint(ws, row, col, fillc, border=True):
    """只上色，不写值（用于合并单元格的从属格）"""
    c = ws.cell(row=row, column=col)
    c.fill = fill(fillc)
    if border:
        c.border = BOX
    return c


def banner(ws, row, col_end, text, sub=None):
    """顶部标题带"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT_BASE, size=13, bold=True, color="FFFFFF")
    c.fill = fill(F_HEADER)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26

def headrow(ws, row, headers, widths=None, height=30, fillc=None):
    for i, h in enumerate(headers, start=1):
        put(ws, row, i, h, font=f_head(), fillc=fillc or F_HEADER, align=AL_C)
    ws.row_dimensions[row].height = height
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

def setw(ws, widths, start=1):
    for i, w in enumerate(widths, start=start):
        ws.column_dimensions[get_column_letter(i)].width = w

CL = get_column_letter
