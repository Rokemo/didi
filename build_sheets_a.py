# -*- coding: utf-8 -*-
"""参数层：07工序节拍 / 08字典 / 09检验标准 / 原始数据表 / 产品资料表 / 发货需求表"""
import re
from openpyxl.utils import get_column_letter as CL
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from build_part1 import *
from build_part1 import _serial_to_date, _to_int
from build_util import *

DATE_FMT = "yyyy-mm-dd"


def _coerce(h, v, date_headers):
    """原始导出里很多数值是文本形态（如 '100.0'），直接 SUMIFS 会被当成 0。
    这里把数值型文本转成真正的数字；日期列转成日期对象；其余保持原样。"""
    if v is None or (isinstance(v, str) and v.strip() in ("", "None", "null", "Null")):
        return None
    if h in date_headers:
        if isinstance(v, (int, float)):
            return _serial_to_date(v)
        if isinstance(v, str):
            try:
                return _serial_to_date(float(v))
            except Exception:
                return v
        return v
    if isinstance(v, str):
        s = v.strip()
        if re.fullmatch(r"-?\d+(\.\d+)?", s):
            return float(s)
    return v


# ============ 07 参数-工序节拍 ============
def build_tact(wb):
    ws = wb.create_sheet(SH_TACT)
    banner(ws, 1, 24, "参数 · 工序节拍库    （修改这里 = 全表排期自动重算）")
    ws.merge_cells("A2:X2")
    put(ws, 2, 1, "偏移口径：计划开工 = 合同日期 + a天 ；最晚完工 = 计划出货日 - b天（表中b以负数记）。"
                  "「NA」= 该类目不走这道工序。黄底行为回落值，需与工厂确认后修正。",
        font=f_note(9, "1F4E79", False), fillc="EAF3FB", align=AL_L)
    ws.row_dimensions[2].height = 30

    put(ws, 3, 1, "标准工艺类目", font=f_head(), fillc=F_SUBHEAD)
    put(ws, 3, 2, "数量档位", font=f_head(), fillc=F_SUBHEAD)
    put(ws, 3, 3, "是否新单", font=f_head(), fillc=F_SUBHEAD)
    put(ws, 3, 4, "查询Key", font=f_head(), fillc=F_SUBHEAD)
    for i, s in enumerate(STAGES):
        c0 = 5 + i*2
        ws.merge_cells(start_row=3, start_column=c0, end_row=3, end_column=c0+1)
        put(ws, 3, c0, "P%d %s" % (i+1, s), font=f_head(), fillc=F_HEADER)
        paint(ws, 3, c0+1, F_HEADER)
    put(ws, 3, 23, "参数来源 / 备注", font=f_head(), fillc=F_SUBHEAD)

    for j, t in enumerate(["标准工艺类目", "数量档位", "是否新单", "查询Key"], start=1):
        put(ws, 4, j, t, font=f_head(9), fillc=F_SUBHEAD)
    for i in range(9):
        c0 = 5 + i*2
        put(ws, 4, c0, "开工 = 合同日+", font=f_head(9), fillc=F_SUBHEAD)
        put(ws, 4, c0+1, "完工 = 出货日", font=f_head(9), fillc=F_SUBHEAD)
    put(ws, 4, 23, "参数来源 / 备注", font=f_head(9), fillc=F_SUBHEAD)
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 30

    rows = tact_rows()
    r = 5
    for cat, tier, nw, offs, note, is_fb in rows:
        bg = F_ASSUME if is_fb else None
        put(ws, r, 1, cat, font=f_input(), fillc=bg, align=AL_C)
        put(ws, r, 2, tier, font=f_input(), fillc=bg)
        put(ws, r, 3, nw, font=f_input(), fillc=bg)
        put(ws, r, 4, "=A%d&\"|\"&B%d&\"|\"&C%d" % (r, r, r), font=f_formula(9), fillc=F_GREY)
        for i, o in enumerate(offs):
            c0 = 5 + i*2
            if o == NA:
                put(ws, r, c0, "NA", font=Font(name=FONT_BASE, size=9, color="A6A6A6"), fillc=bg or F_GREY)
                put(ws, r, c0+1, "NA", font=Font(name=FONT_BASE, size=9, color="A6A6A6"), fillc=bg or F_GREY)
            else:
                put(ws, r, c0, o[0], font=f_input(), fillc=bg, numfmt="+0;-0;0")
                put(ws, r, c0+1, o[1], font=f_input(), fillc=bg, numfmt="+0;-0;0")
        put(ws, r, 23, note, font=f_note(9, "C55A11" if is_fb else "808080", is_fb), fillc=bg, align=AL_L)
        r += 1

    setw(ws, [15, 11, 9, 26] + [11, 11]*9 + [46])
    ws.freeze_panes = "E5"
    ws.auto_filter.ref = "A4:W%d" % (r-1)
    return r-1


# ============ 08 参数-字典（数据驱动） ============
def build_dict(wb, mats, facs, buyers):
    ws = wb.create_sheet(SH_DICT)
    banner(ws, 1, 11, "参数 · 字典与阈值    （新增工厂/品类/调整预警天数，都改这里）")

    # --- 区块1 标准工艺类目总表（类目取自「产品资料表」的 产品编号→工艺品类，不再系统推断）
    ws.merge_cells("A3:C3")
    put(ws, 3, 1, "① 标准工艺类目总表（类目不再由系统推断，统一取自「产品资料表」的 产品编号→工艺品类；维护产品资料表即可）",
        font=f_title(11), fillc=F_KPI, align=AL_L)
    for j, t in enumerate(["标准工艺类目", "是否纳入节拍库", "说明"], start=1):
        put(ws, 4, j, t, font=f_head(), fillc=F_SUBHEAD)
    r = 5
    for cat in CAT_ORDER:
        put(ws, r, 1, cat, font=f_input(bold=True))
        put(ws, r, 2, "是（07参数有节拍）", font=f_input())
        put(ws, r, 3, "产品资料表「工艺品类」填写此值，02 即按此匹配 07 工序节拍", font=f_note(9, "808080", False), align=AL_L)
        r += 1
    put(ws, r, 1, "（其他/待确认）", font=f_input())
    put(ws, r, 2, "否", font=f_input())
    put(ws, r, 3, "工艺品类留空 或 产品编号不在产品资料表 → 02 显示「待确认」，需在产品资料表补全", font=f_note(9, "C00000", False), align=AL_L)
    r += 1
    map_end = r - 1

    # --- 区块2 工序名映射
    r += 1
    hd2 = r
    ws.merge_cells(start_row=hd2, start_column=1, end_row=hd2, end_column=10)
    put(ws, hd2, 1, "② 各类目工序名称对照（「—」= 该类目不走此工序，排期表自动跳过）",
        font=f_title(11), fillc=F_KPI, align=AL_L)
    put(ws, hd2+1, 1, "标准工艺类目", font=f_head(), fillc=F_SUBHEAD)
    for i, s in enumerate(STAGES):
        put(ws, hd2+1, 2+i, "P%d\n%s" % (i+1, s), font=f_head(9), fillc=F_SUBHEAD)
    ws.row_dimensions[hd2+1].height = 32
    r = hd2 + 2
    stage_start = r
    for cat in CAT_ORDER:
        put(ws, r, 1, cat, font=f_input(bold=True))
        for i, nm in enumerate(CAT_STAGE_NAME[cat]):
            put(ws, r, 2+i, nm, font=f_input() if nm != "—" else Font(name=FONT_BASE, size=10, color="BFBFBF"))
        r += 1
    stage_end = r - 1

    # --- 区块3 阈值
    r += 1
    hd3 = r
    ws.merge_cells(start_row=hd3, start_column=1, end_row=hd3, end_column=4)
    put(ws, hd3, 1, "③ 交期预警阈值（黄底=可调假设项）", font=f_title(11), fillc=F_KPI, align=AL_L)
    for j, t in enumerate(["参数名", "取值", "单位", "用途说明"], start=1):
        put(ws, hd3+1, j, t, font=f_head(), fillc=F_SUBHEAD)
    thr = [
        ("紧急阈值", 7, "天", "距计划出货 ≤ 该天数 且未交清 → 标记「紧急」"),
        ("预警阈值", 15, "天", "距计划出货 ≤ 该天数 且未交清 → 标记「预警」"),
        ("呆滞阈值", 180, "天", "逾期超过该天数 → 标记「呆滞」，进入专项清理"),
        ("紧急低完成率", 0.8, "比例", "紧急期内完成率低于此值 → 风险升为「高」"),
        ("预警低完成率", 0.5, "比例", "预警期内完成率低于此值 → 风险升为「中」"),
    ]
    r = hd3 + 2
    thr_start = r
    for name, val, unit, note in thr:
        put(ws, r, 1, name, font=f_formula(bold=True), fillc=F_ASSUME)
        put(ws, r, 2, val, font=f_input(bold=True), fillc=F_ASSUME,
            numfmt="0%" if isinstance(val, float) else "0")
        put(ws, r, 3, unit, font=f_note(9, "808080", False), fillc=F_ASSUME)
        put(ws, r, 4, note, font=f_note(9), fillc=F_ASSUME, align=AL_L)
        r += 1

    # --- 区块4 下拉选项源（由数据去重生成）
    r += 1
    hd4 = r
    ws.merge_cells(start_row=hd4, start_column=1, end_row=hd4, end_column=6)
    put(ws, hd4, 1, "④ 下拉选项源（新增工厂/跟单员请在此列往下追加）", font=f_title(11), fillc=F_KPI, align=AL_L)
    lists = {
        "工厂": sorted(f for f in facs if f),
        "跟单员": sorted(b for b in buyers if b),
        "原始品类": sorted(m if m else "(空)" for m in mats),
        "是否新单": ["否", "是"],
        "跟进结论": DEFAULT_QC_LIST,
    }
    for j, (name, vals) in enumerate(lists.items(), start=1):
        put(ws, hd4+1, j, name, font=f_head(), fillc=F_SUBHEAD)
        for k, v in enumerate(vals):
            put(ws, hd4+2+k, j, v, font=f_input())
    list_start = hd4 + 2
    list_rows = max(len(v) for v in lists.values())

    setw(ws, [22, 16, 40, 16, 16, 18, 14, 14, 14, 16])
    ws.column_dimensions["C"].width = 52

    return dict(map_start=5, map_end=map_end,
                stage_start=stage_start, stage_end=stage_end,
                thr_start=thr_start,
                list_start=list_start, list_end=list_start+list_rows-1,
                lists=lists)


# ============ 09 检验标准 ============
def build_qc(wb, qc):
    ws = wb.create_sheet(SH_QC)
    banner(ws, 1, 8, "产中检验时间标准（原 Sheet1 结构化整理）")
    ws.merge_cells("A2:H2")
    put(ws, 2, 1, "用法：③工序进度 中某道工序进入「进行中」时，对照本表安排验货；"
                  "老款请优先抽查客诉数据 TOP3 的问题项，并与工厂技术共同确认整改措施。",
        font=f_note(9, "1F4E79", False), fillc="EAF3FB", align=AL_L)
    ws.row_dimensions[2].height = 26
    hd = ["序号", "问题大类", "问题小类", "关联工序", "合理验货时间点", "检验标准/方法", "责任人", "备注"]
    for j, t in enumerate(hd, start=1):
        put(ws, 3, j, t, font=f_head(), fillc=F_HEADER)
    ws.row_dimensions[3].height = 28
    r = 4
    for i, v in enumerate(qc, start=1):
        put(ws, r, 1, i, font=f_formula(9))
        for j in range(7):
            put(ws, r, 2+j, v[j], font=f_input(9), align=AL_L)
        r += 1
    setw(ws, [6, 14, 34, 18, 16, 32, 12, 40])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:H%d" % (r-1)


# ============ 原始数据表（主数据源） ============
def build_raw(wb, raw_rows):
    ws = wb.create_sheet(SH_RAW)
    ncols = len(raw_rows[0]) if raw_rows else 0
    banner(ws, 1, min(20, ncols), "原始数据表 · 主数据源（从 D:\\原始数据.xlsx 导入 %d 行；"
          "左侧 1~%d 列粘贴ERP导出，右侧「手工维护」列由你填写，粘贴导出时勿覆盖" % (len(raw_rows), ncols))
    headers = list(raw_rows[0].keys())
    # 表头（第2行）：导出列 + 手工维护列
    manual = [("最新跟进日期", "手工维护·填日期"), ("跟进结论", "手工维护·下拉"),
              ("跟进备注", "手工维护·自由文本"), ("是否新单", "手工维护·默认否")]
    all_hd = headers + [m[0] + "(%s)" % m[1] for m in manual]
    for j, h in enumerate(all_hd, start=1):
        is_man = j > ncols
        put(ws, 2, j, h, font=f_head(9), fillc="C55A11" if is_man else "808080", align=AL_L)
    ws.row_dimensions[2].height = 30

    date_headers = set(h for h in headers if ("日期" in h or "时间" in h))
    r = 3
    for row in raw_rows:
        for j, h in enumerate(headers, start=1):
            v = row.get(h)
            v = _coerce(h, v, date_headers)
            put(ws, r, j, v, font=f_input(9), align=AL_L if j in (1, 2, 3) else AL_C)
        # 手工维护列：前3空，是否新单默认否
        put(ws, r, ncols+1, None, font=f_input(9), fillc="FFF0E1", numfmt=DATE_FMT)
        put(ws, r, ncols+2, None, font=f_input(9), fillc="FFF0E1")
        put(ws, r, ncols+3, None, font=f_input(9), fillc="FFF0E1", align=AL_L)
        put(ws, r, ncols+4, "否", font=f_input(9), fillc="FFF0E1")
        r += 1
    last = r - 1

    # 数据验证：是否新单 / 跟进结论（仅手工维护列）
    dv1 = DataValidation(type="list", formula1='"否,是"', allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv1); dv1.add("%s3:%s%d" % (CL(ncols+4), CL(ncols+4), last))
    dv2 = DataValidation(type="list", formula1='"%s"' % ",".join(DEFAULT_QC_LIST), allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv2); dv2.add("%s3:%s%d" % (CL(ncols+2), CL(ncols+2), last))

    setw(ws, [12]*ncols + [13, 14, 30, 10])
    ws.freeze_panes = "D3"
    ws.auto_filter.ref = "A2:%s%d" % (CL(len(all_hd)), last)
    ws.sheet_view.showGridLines = False


# ============ 产品资料表（用户维护：产品编号→工艺品类） ============
def build_prod(wb, prod_rows):
    ws = wb.create_sheet(SH_PROD)
    n = len(prod_rows)
    banner(ws, 1, 4, "产品资料表 · 类目来源（从 D:\\产品资料模板.xlsx 导入 %d 行；你维护，"
          "02 按产品编号 VLOOKUP 取「工艺品类」）" % n)
    ws.merge_cells("A2:C2")
    put(ws, 2, 1, "用法：①列=产品编号（须与原始数据表一致）；③列=工艺品类（标准工艺类目口径，可下拉选）。"
                  "留空则该产品编号在 02/03 的类目显示「待确认」。黄底为可填项。",
        font=f_note(9, "1F4E79", False), fillc="EAF3FB", align=AL_L)
    ws.row_dimensions[2].height = 28
    for j, t in enumerate(["产品编号", "中文品名", "工艺品类"], start=1):
        put(ws, 3, j, t, font=f_head(9), fillc="C55A11" if j == 3 else "808080", align=AL_L)
    r = 4
    for row in prod_rows:
        sku = (row.get("产品编号") or "").strip()
        nm = (row.get("中文品名") or "").strip()
        cat = row.get("工艺品类")
        if cat in (None, "None", "null", ""):
            cat = None
        put(ws, r, 1, sku, font=f_input(9), align=AL_L)
        put(ws, r, 2, nm, font=f_input(9), align=AL_L)
        put(ws, r, 3, cat, font=f_input(9, bold=True), fillc="FFF2CC", align=AL_L)
        r += 1
    last = r - 1
    # 数据验证：工艺品类 下拉（标准工艺类目口径）
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(CAT_ORDER), allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv); dv.add("C4:C%d" % last)
    setw(ws, [22, 40, 18])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:C%d" % last
    ws.sheet_view.showGridLines = False


# ============ 发货需求表（用户维护） ============
def build_ship(wb, ship_rows):
    ws = wb.create_sheet(SH_SHIP)
    n = len(ship_rows)
    banner(ws, 1, 7, "发货需求表 · 用户维护（从 D:\\发货需求模板.xlsx 导入 %d 行；"
          "「10发货需求判断」按 产品编号+团队 汇总对比未入库数量）" % n)
    ws.merge_cells("A2:G2")
    put(ws, 2, 1, "用法：每行一条发货需求。①SKU=产品编号；②团队（如 鄢紫薇/UP-US，须与原始数据表团队一致）；"
                  "⑥发货数量；⑦发货时间（日期）。在「10发货需求判断」设发货起止日即可看窗口内需求能否被未入库数量满足。",
        font=f_note(9, "1F4E79", False), fillc="EAF3FB", align=AL_L)
    ws.row_dimensions[2].height = 28
    hd = ["SKU", "团队", "运营", "工厂", "跟单", "发货数量", "发货时间"]
    for j, t in enumerate(hd, start=1):
        put(ws, 3, j, t, font=f_head(9), fillc="C55A11" if j in (1, 2, 6, 7) else "808080", align=AL_L)
    r = 4
    for row in ship_rows:
        sku = (row.get("SKU") or "").strip()
        team = (row.get("团队") or "").strip()
        op = (row.get("运营") or "").strip()
        fac = (row.get("工厂") or "").strip()
        trk = (row.get("跟单") or "").strip()
        qty = _to_num(row.get("发货数量"))
        dt = _serial_to_date(_to_num(row.get("发货时间")))
        put(ws, r, 1, sku, font=f_input(9), align=AL_L)
        put(ws, r, 2, team, font=f_input(9), align=AL_L)
        put(ws, r, 3, op, font=f_input(9), align=AL_L)
        put(ws, r, 4, fac, font=f_input(9), align=AL_L)
        put(ws, r, 5, trk, font=f_input(9), align=AL_L)
        put(ws, r, 6, qty, font=f_input(9), align=AL_C, numfmt="#,##0")
        put(ws, r, 7, dt, font=f_input(9), align=AL_C, numfmt=DATE_FMT)
        r += 1
    last = r - 1
    setw(ws, [22, 18, 12, 14, 12, 12, 12])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:G%d" % last
    ws.sheet_view.showGridLines = False


def _to_num(v):
    if v is None or (isinstance(v, str) and v.strip() in ("", "None", "null")):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except Exception:
            return None
    return None


# ============ 销售预测表（用户维护：产品编号+运营专员 → 未来3月预计出货数量） ============
FCST_NB = 500   # 显示行缓冲（可继续往下追加新组合）


def build_fcst(wb, combos, name_map):
    """combos: [(产品编号, 运营专员), ...] 按冗余金额降序；name_map: 产品编号->中文品名"""
    ws = wb.create_sheet(SH_FCST)
    n = len(combos)
    banner(ws, 1, 6, "销售预测表 · 用户维护（已按当前在手数据预生成 %d 个「产品编号 × 运营专员」组合，"
          "按待交付金额从大到小排；只需填 D 列）" % n)
    ws.merge_cells("A2:F2")
    put(ws, 2, 1, "用途：「11合同冗余」的接单判断依赖此表。D列「未来3月预计出货数量」是唯一需要你填的（黄底）。"
                  "填了才会算出 S/A/B/C 等级、安全库存和可否接单；留空则该行显示「待填预测」。"
                  "分级口径：>500=S(核心快销,安全库存=月均×1.5) / >200=A(畅销,×1.0) / >50=B(平销,×0.8) / ≤50=C(滞销,×0.5)。",
        font=f_note(9, "1F4E79", False), fillc="EAF3FB", align=AL_L)
    ws.row_dimensions[2].height = 42

    hd = [("产品编号", 22), ("运营专员", 12), ("组合键(自动)", 26),
          ("未来3月预计出货数量", 18), ("中文品名(自动)", 34), ("当前在手待交付(自动)", 16)]
    for j, (t, w) in enumerate(hd, start=1):
        put(ws, 3, j, t, font=f_head(9), fillc="C55A11" if j == 4 else "808080", align=AL_L)
        ws.column_dimensions[CL(j)].width = w
    ws.row_dimensions[3].height = 30

    RSKU = "%s!$C$3:$C$602" % q(SH_RAW)
    ROP = "%s!$E$3:$E$602" % q(SH_RAW)
    RQTY = "%s!$AG$3:$AG$602" % q(SH_RAW)     # 未到货总数量
    PROD = "%s!$A$4:$C$2000" % q(SH_PROD)

    for k in range(FCST_NB):
        r = 4 + k
        sku, op = combos[k] if k < n else ("", "")
        put(ws, r, 1, sku or None, font=f_input(9), align=AL_L)
        put(ws, r, 2, op or None, font=f_input(9), align=AL_L)
        put(ws, r, 3, '=IF($A%d="","",$A%d&"|"&$B%d)' % (r, r, r), font=f_formula(9), fillc=F_GREY, align=AL_L)
        put(ws, r, 4, None, font=f_input(9, bold=True), fillc=F_ASSUME, align=AL_C, numfmt="#,##0")
        put(ws, r, 5, '=IF($A%d="","",IFERROR(VLOOKUP($A%d,%s,2,FALSE),""))' % (r, r, PROD),
            font=f_xref(9), align=AL_L)
        put(ws, r, 6, '=IF($A%d="","",SUMIFS(%s,%s,$A%d,%s,$B%d))' % (r, RQTY, RSKU, r, ROP, r),
            font=f_formula(9), align=AL_C, numfmt="#,##0")

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = "A3:F%d" % (3 + FCST_NB)
    ws.sheet_view.showGridLines = False
    return n
