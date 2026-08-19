# -*- coding: utf-8 -*-
"""
跟单进度网页原型 - 零依赖本地服务
技术栈：Python 标准库 http.server + sqlite3
启动：python server.py  然后浏览器打开 http://127.0.0.1:8731

v0.2 新增：
  - 登录+权限（角色 admin / 跟单员 / viewer）
  - 字段校验（后端权威校验，逐行返回错误）
  - 批量粘贴导入（POST /api/raw/bulk）
  - 跟单员/运营组别/厂商 维度筛选
  - 核心逻辑：逾期呆滞（/api/overdue）、合同冗余判定（/api/redundancy）
  - v0.3：原始数据表支持勾选+批量删除（DELETE /api/raw/batch）、Excel 文件导入（POST /api/raw/import）
  - v0.4：批量删除支持「跨页全选所有匹配项」（DELETE /api/raw/batch?scope=all，按当前搜索/筛选条件整批删除）
"""
import sqlite3
import json
import os
import datetime
import shutil
import threading
import hashlib
import secrets
import time
import re
import io
import openpyxl
import proj_logic as PL
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs, quote

BASE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE, "tracker.db")
BACKUP_DIR = os.path.join(BASE, "backups")
STATIC = os.path.join(BASE, "static")
PORT = int(os.environ.get("PORT", "8731"))
HOST = os.environ.get("HOST", "127.0.0.1")

# 鉴权开关：默认关闭（免登录，所有人按管理员处理）。
# 需要恢复账号密码时，启动前设置环境变量 AUTH_ENABLED=1 即可。
AUTH_ENABLED = os.environ.get("AUTH_ENABLED", "0") in ("1", "true", "True")

SCHEMA = """
CREATE TABLE IF NOT EXISTS raw_orders (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  contract_no TEXT,
  product_code TEXT,
  team TEXT,
  operator TEXT,
  operator_group TEXT,
  factory TEXT,
  order_qty REAL,
  delivered_qty REAL,
  undelivered_qty REAL,
  outstanding_qty REAL,
  unit_price REAL,
  delivery_date TEXT,
  customs_material TEXT,
  followup_date TEXT,
  followup_conclusion TEXT,
  followup_note TEXT,
  is_new TEXT,
  created_at TEXT DEFAULT (datetime('now','localtime')),
  updated_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS products (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  product_code TEXT UNIQUE,
  product_name TEXT,
  craft_category TEXT,
  spu TEXT
);
CREATE TABLE IF NOT EXISTS shipping_demands (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  sku TEXT,
  team TEXT,
  operator TEXT,
  factory TEXT,
  follower TEXT,
  ship_qty REAL,
  ship_date TEXT
);
CREATE TABLE IF NOT EXISTS sales_forecast (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  product_code TEXT,
  operator TEXT,
  forecast_qty REAL
);
CREATE TABLE IF NOT EXISTS users (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  username TEXT UNIQUE,
  password_hash TEXT,
  role TEXT,
  display_name TEXT
);
CREATE TABLE IF NOT EXISTS proc_stages (
  pkey TEXT NOT NULL,
  stage_idx INTEGER NOT NULL,
  actual_date TEXT,
  PRIMARY KEY (pkey, stage_idx)
);
CREATE TABLE IF NOT EXISTS proc_detail (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  contract_no TEXT,
  product_code TEXT,
  spu TEXT,
  order_qty REAL,
  contract_date TEXT,
  ship_date TEXT,
  craft_category TEXT,
  import_batch TEXT,
  created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS proc_meta (
  product_code TEXT NOT NULL,
  contract_no TEXT NOT NULL,
  customer_complaint TEXT,
  inspection_focus TEXT,
  inspection_time TEXT,
  inspection_result TEXT,
  PRIMARY KEY (product_code, contract_no)
);
"""

PROD_COLS = ["product_code", "product_name", "craft_category", "spu"]
SHIP_COLS = ["sku", "team", "operator", "factory", "follower", "ship_qty", "ship_date"]
FC_COLS = ["product_code", "operator", "forecast_qty"]

# ---------------- 三张可导入表的表头映射/必填/配置 ----------------
PROD_ALIASES = {
    "product_code": ["产品编号", "产品编码", "货号"],
    "product_name": ["产品品名", "产品名称", "品名", "中文品名"],
    "craft_category": ["工艺品类", "标准工艺类目", "类目"],
    "spu": ["SPU", "spu"],
}
SHIP_ALIASES = {
    "sku": ["产品编号", "SKU", "产品编码"],
    "team": ["团队"],
    "operator": ["运营专员", "运营"],
    "factory": ["厂商简称", "厂商"],
    "follower": ["跟单员", "跟单"],
    "ship_qty": ["发货数量", "发货量"],
    "ship_date": ["发货时间", "发货日期"],
}
FC_ALIASES = {
    "product_code": ["产品编号", "产品编码", "货号"],
    "operator": ["运营专员", "运营"],
    "forecast_qty": ["预计出货数量", "预测数量", "出货数量", "未来3月预计出货数量"],
}
# key -> 表名/列/中文表头/必填列/别名。* 表示必填。
IMPORT_CFG = {
    "products": {
        "table": "products", "cols": PROD_COLS, "aliases": PROD_ALIASES,
        "labels": {"product_code": "产品编号*", "product_name": "中文品名", "craft_category": "工艺品类", "spu": "SPU"},
        "required": ["product_code"],
    },
    "shipping": {
        "table": "shipping_demands", "cols": SHIP_COLS, "aliases": SHIP_ALIASES,
        "labels": {"sku": "产品编号(SKU)*", "team": "团队*", "operator": "运营专员",
                   "factory": "厂商", "follower": "跟单员", "ship_qty": "发货数量*",
                   "ship_date": "发货时间(YYYY-MM-DD)*"},
        "required": ["sku", "ship_qty", "ship_date"],
    },
    "forecast": {
        "table": "sales_forecast", "cols": FC_COLS, "aliases": FC_ALIASES,
        "labels": {"product_code": "产品编号*", "operator": "运营专员*", "forecast_qty": "未来3月预计出货数量*"},
        "required": ["product_code", "operator", "forecast_qty"],
    },
}

# ---------------- 工序进度（独立数据源，与 raw_orders 解耦） ----------------
# 导入明细列：合同号 + 产品编号 + 订单数量 + 合同日期 + 出货日期 + 可选 SPU/工艺品类
PROC_COLS = ["contract_no", "product_code", "spu", "order_qty", "contract_date", "ship_date", "craft_category"]
PROC_ALIASES = {
    "contract_no": ["合同号", "合同编号"],
    "product_code": ["产品编号", "产品编码", "货号", "SKU"],
    "spu": ["SPU", "spu"],
    "order_qty": ["订单数量", "订单总数量", "数量"],
    "contract_date": ["合同日期"],
    "ship_date": ["出货日期", "交货日期", "船期"],
    "craft_category": ["工艺品类", "标准工艺类目", "类目"],
}
PROC_LABELS = {
    "contract_no": "合同号*",
    "product_code": "产品编号*",
    "order_qty": "订单数量*",
    "contract_date": "合同日期(YYYY-MM-DD)",
    "ship_date": "出货日期(YYYY-MM-DD)",
    "spu": "SPU",
    "craft_category": "工艺品类",
}
PROC_REQUIRED = ["contract_no", "product_code", "order_qty"]
# 检验重点下拉选项（16 项）
INSPECTION_FOCUS = [
    "金加工阶段", "外发返厂", "半成品成型", "半成品备件", "预组装阶段", "入箱前",
    "原材料到厂阶段", "产前样", "喷塑环节", "来料检查", "打磨/喷塑检验", "预包装阶段",
    "样品开发阶段", "下料后", "预加工", "产前样组装测试",
]
PROC_FILTER_COLS = ["contract_no", "product_code", "spu", "craft_category", "inspection_focus"]

RAW_COLS = [
    "contract_no", "product_code", "team", "operator", "operator_group",
    "factory", "order_qty", "delivered_qty", "undelivered_qty",
    "outstanding_qty", "unit_price", "contract_date", "delivery_date",
    "customs_material",
    "followup_date", "followup_conclusion", "followup_note", "is_new"
]
RAW_LABELS = {
    "contract_no": "合同号", "product_code": "产品编号", "team": "团队",
    "operator": "运营专员", "operator_group": "运营组别", "factory": "厂商",
    "order_qty": "订单总数量", "delivered_qty": "已交付数量",
    "undelivered_qty": "未入库数量", "outstanding_qty": "未到货总数量",
    "unit_price": "采购单价", "contract_date": "合同日期", "delivery_date": "交货日期",
    "customs_material": "报关材质", "followup_date": "最新跟进日期",
    "followup_conclusion": "跟进结论", "followup_note": "跟进备注",
    "is_new": "是否新单"
}
NUMCOLS = {"order_qty", "delivered_qty", "undelivered_qty", "outstanding_qty", "unit_price"}
# 跟单员维度筛选允许的白名单列
FILTER_COLS = ["operator", "operator_group", "factory"]


def build_raw_where(search, flt):
    """构造 raw_orders 的筛选 WHERE 子句与参数（GET 列表与跨页批量删除共用）。

    search: 关键字（按 合同号/产品编号/团队/运营/厂商/运营组别 模糊匹配）
    flt:    {col: value} 精确筛选（来自 FILTER_COLS）
    返回 (where, args)
    """
    where = ""
    args = []
    if search:
        where = "WHERE " + " OR ".join(["%s LIKE ?" % col for col in
                  ["contract_no", "product_code", "team", "operator",
                   "factory", "operator_group"]])
        args = ["%" + search + "%"] * 6
    if flt:
        clause = " AND ".join(["%s=?" % col for col in flt])
        where = ("WHERE " + clause) if not where else (where + " AND " + clause)
        args += [flt[col] for col in flt]
    return where, args

# 密码/会话
SESSIONS = {}  # token -> {username, role, display_name, exp}
SESSION_TTL = 86400


def conn_db():
    c = sqlite3.connect(DB)
    c.row_factory = sqlite3.Row
    return c


def init_db():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    c = conn_db()
    c.executescript(SCHEMA)
    # 迁移：老数据库可能缺 contract_date 列
    cols = [r[1] for r in c.execute("PRAGMA table_info(raw_orders)")]
    if "contract_date" not in cols:
        c.execute("ALTER TABLE raw_orders ADD COLUMN contract_date TEXT")
    # 迁移：产品资料补 SPU 列
    pcols = [r[1] for r in c.execute("PRAGMA table_info(products)")]
    if "spu" not in pcols:
        c.execute("ALTER TABLE products ADD COLUMN spu TEXT")
    # 迁移：工序进度解耦 raw_orders —— proc_stages 由 raw_id 主键改为复合键 pkey
    scols = [r[1] for r in c.execute("PRAGMA table_info(proc_stages)")]
    if "raw_id" in scols:
        # raw_orders 已清空，旧 proc_stages 不再有有效关联，直接重建
        c.execute("DROP TABLE proc_stages")
        c.execute("""CREATE TABLE proc_stages (
            pkey TEXT NOT NULL, stage_idx INTEGER NOT NULL, actual_date TEXT,
            PRIMARY KEY (pkey, stage_idx))""")
    seed_users(c)
    c.commit()
    c.close()


def hash_pw(pw, salt=None):
    if salt is None:
        salt = secrets.token_hex(8)
    h = hashlib.sha256((salt + pw).encode("utf-8")).hexdigest()
    return salt + ":" + h


def check_pw(pw, stored):
    try:
        salt, h = stored.split(":")
    except Exception:
        return False
    return hash_pw(pw, salt) == stored


def seed_users(c):
    defaults = [
        ("admin", "1fxr3S0A6Jw5DsWE_uM", "admin", "管理员"),
        ("zhang", "tFrzcD_KcD5U1KEciyQ", "跟单员", "张敏(跟单员)"),
        ("view", "bCqOT_aRsggMr48zTkI", "viewer", "访客(只读)"),
    ]
    for u, pw, role, name in defaults:
        if not c.execute("SELECT 1 FROM users WHERE username=?", (u,)).fetchone():
            c.execute("INSERT INTO users (username,password_hash,role,display_name) VALUES (?,?,?,?)",
                      (u, hash_pw(pw), role, name))


# ---------------- 备份 ----------------
def backup_db():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = os.path.join(BACKUP_DIR, "tracker_%s.db" % ts)
    shutil.copy2(DB, dst)
    files = sorted([f for f in os.listdir(BACKUP_DIR) if f.startswith("tracker_") and f.endswith(".db")])
    for old in files[:-30]:
        try:
            os.remove(os.path.join(BACKUP_DIR, old))
        except OSError:
            pass
    return dst


def today_str():
    return datetime.date.today().isoformat()


def to_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s == "":
        return None
    try:
        return float(s)
    except Exception:
        return None


def to_str(v):
    if v is None:
        return ""
    return str(v).strip()


def parse_date(v):
    if v is None:
        return ""
    s = str(v).strip()
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    if len(s) >= 10 and s[4] == "/" and s[7] == "/":
        return s[:10].replace("/", "-")
    return s


def clean_raw(d):
    out = {}
    for c in RAW_COLS:
        v = d.get(c)
        if c in NUMCOLS:
            out[c] = to_num(v)
        elif c == "delivery_date":
            out[c] = parse_date(v)
        else:
            out[c] = to_str(v)
    return out


# ---------------- 校验 ----------------
def _num_ok(v):
    if v is None or v == "":
        return True
    if isinstance(v, (int, float)):
        return True
    s = str(v).strip().replace(",", "")
    if s == "":
        return True
    try:
        float(s)
        return True
    except Exception:
        return False


def validate_raw(d):
    """返回错误字符串列表（空列表=通过）。直接校验原始输入（含字符串），
    避免 clean_raw 把非法数字静默转成 None 导致漏判。"""
    errs = []
    if not to_str(d.get("contract_no")):
        errs.append("合同号 不能为空")
    if not to_str(d.get("product_code")):
        errs.append("产品编号 不能为空")
    for col in ("order_qty", "delivered_qty", "undelivered_qty", "outstanding_qty", "unit_price"):
        v = d.get(col)
        if v not in (None, ""):
            if not _num_ok(v):
                errs.append("%s 必须是数字" % RAW_LABELS.get(col, col))
            elif to_num(v) < 0:
                errs.append("%s 不能为负数" % RAW_LABELS.get(col, col))
    dd = d.get("delivery_date")
    if dd not in (None, ""):
        s = parse_date(dd)
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", s):
            errs.append("交货日期 格式应为 YYYY-MM-DD")
        else:
            try:
                datetime.date.fromisoformat(s)
            except Exception:
                errs.append("交货日期 不是合法日期")
    fd = d.get("followup_date")
    if fd not in (None, ""):
        s = parse_date(fd)
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", s):
            errs.append("最新跟进日期 格式应为 YYYY-MM-DD")
    return errs


def validate_generic(d, cols, required):
    errs = []
    for c in required:
        if not to_str(d.get(c)):
            errs.append("%s 不能为空" % c)
    for c in cols:
        if c in NUMCOLS or c.endswith("qty") or c.endswith("price"):
            v = d.get(c)
            if v not in (None, ""):
                if to_num(v) is None:
                    errs.append("%s 必须是数字" % c)
    if "ship_date" in cols and to_str(d.get("ship_date")):
        s = parse_date(d.get("ship_date"))
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", s):
            errs.append("发货时间 格式应为 YYYY-MM-DD")
    return errs


def validate_proc(obj):
    """工序进度明细导入校验。"""
    e = []
    if not to_str(obj.get("contract_no")):
        e.append("合同号 不能为空")
    if not to_str(obj.get("product_code")):
        e.append("产品编号 不能为空")
    oq = to_str(obj.get("order_qty"))
    if oq not in (None, ""):
        if to_num(oq) is None:
            e.append("订单数量 必须是数字")
    for dc in ("contract_date", "ship_date"):
        v = to_str(obj.get(dc))
        if v and not re.match(r"^\d{4}-\d{2}-\d{2}$", v):
            e.append("%s 格式应为 YYYY-MM-DD" % dc)
    return e


def clean_proc(obj, batch=""):
    """工序进度明细清洗：文本字段转字符串，订单数量转数字。"""
    d = {}
    for col in PROC_COLS:
        d[col] = to_str(obj.get(col)) if obj.get(col) is not None else None
    try:
        d["order_qty"] = float(d["order_qty"]) if d["order_qty"] not in (None, "") else 0.0
    except Exception:
        d["order_qty"] = 0.0
    if batch:
        d["import_batch"] = batch
    return d


# ---------------- KPI ----------------
def compute_kpi():
    c = conn_db()
    t = today_str()
    row = c.execute("""
        SELECT
          COUNT(*) AS rows,
          COUNT(DISTINCT contract_no) AS contracts,
          COUNT(DISTINCT factory) AS factories,
          COUNT(DISTINCT product_code) AS products,
          COALESCE(SUM(order_qty),0) AS order_qty,
          COALESCE(SUM(delivered_qty),0) AS delivered_qty,
          COALESCE(SUM(outstanding_qty),0) AS outstanding_qty,
          COALESCE(SUM(outstanding_qty*unit_price),0) AS outstanding_amt
        FROM raw_orders
    """).fetchone()
    overdue = c.execute("""
        SELECT COUNT(*) AS n FROM raw_orders
        WHERE delivery_date IS NOT NULL AND delivery_date <> ''
          AND delivery_date < ? AND COALESCE(outstanding_qty,0) > 0
    """, (t,)).fetchone()["n"]
    fac = c.execute("""
        SELECT factory AS name, COALESCE(SUM(outstanding_qty*unit_price),0) AS amt
        FROM raw_orders GROUP BY factory ORDER BY amt DESC LIMIT 10
    """).fetchall()
    grp = c.execute("""
        SELECT operator_group AS name, COUNT(*) AS n FROM raw_orders
        WHERE delivery_date IS NOT NULL AND delivery_date <> ''
          AND delivery_date < ? AND COALESCE(outstanding_qty,0) > 0
        GROUP BY operator_group
    """, (t,)).fetchall()
    c.close()
    order_qty = row["order_qty"] or 0
    delivered = row["delivered_qty"] or 0
    rate = (delivered / order_qty * 100) if order_qty else 0
    return {
        "rows": row["rows"], "contracts": row["contracts"],
        "factories": row["factories"], "products": row["products"],
        "order_qty": round(order_qty, 0), "delivered_qty": round(delivered, 0),
        "outstanding_qty": round(row["outstanding_qty"] or 0, 0),
        "outstanding_amt": round(row["outstanding_amt"] or 0, 2),
        "overdue": overdue, "completion_rate": round(rate, 1),
        "by_factory": [dict(r) for r in fac],
        "by_group": [dict(r) for r in grp],
    }


# ---------------- 逾期呆滞 ----------------
OVERDUE_BUCKETS = [("0-30", 0, 30), ("31-90", 31, 90), ("91-180", 91, 180),
                   ("181-365", 181, 365), ("365+", 366, 10**9)]


def bucket_of(days):
    for name, lo, hi in OVERDUE_BUCKETS:
        if lo <= days <= hi:
            return name
    return "0-30"


def compute_overdue(base, group="", operator="", factory="", bucket="", page=1, size=200):
    c = conn_db()
    rows = c.execute("""SELECT id, contract_no, product_code, operator, operator_group,
        factory, outstanding_qty, unit_price, delivery_date, followup_date
        FROM raw_orders""").fetchall()
    c.close()
    try:
        base_d = datetime.date.fromisoformat(base)
    except Exception:
        base_d = datetime.date.today()
    summary = {b: {"count": 0, "amt": 0.0} for b, _, _ in OVERDUE_BUCKETS}
    stale_count = 0
    overdue_count = 0
    overdue_amt = 0.0
    matched = []
    for r in rows:
        dd = r["delivery_date"]
        if not dd:
            continue
        try:
            dd_d = datetime.date.fromisoformat(parse_date(dd))
        except Exception:
            continue
        oq = r["outstanding_qty"] or 0
        if oq <= 0:
            continue
        if dd_d >= base_d:
            continue  # 未到期
        # 逾期
        days = (base_d - dd_d).days
        amt = oq * (r["unit_price"] or 0)
        overdue_count += 1
        overdue_amt += amt
        bname = bucket_of(days)
        summary[bname]["count"] += 1
        summary[bname]["amt"] += amt
        # 呆滞：超 90 天无跟进
        fd = r["followup_date"]
        stale = False
        if not fd:
            stale = True
        else:
            try:
                fd_d = datetime.date.fromisoformat(parse_date(fd))
                if (base_d - fd_d).days > 90:
                    stale = True
            except Exception:
                stale = True
        if stale:
            stale_count += 1
        # 过滤
        if group and r["operator_group"] != group:
            continue
        if operator and r["operator"] != operator:
            continue
        if factory and r["factory"] != factory:
            continue
        if bucket and bname != bucket:
            continue
        matched.append({
            "id": r["id"], "contract_no": r["contract_no"], "product_code": r["product_code"],
            "operator": r["operator"], "operator_group": r["operator_group"], "factory": r["factory"],
            "outstanding_qty": round(oq, 0), "amt": round(amt, 2), "days": days,
            "bucket": bname, "stale": stale, "followup_date": fd or "",
        })
    matched.sort(key=lambda x: -x["days"])
    total = len(matched)
    start = (page - 1) * size
    page_rows = matched[start:start + size]
    return {
        "base": base,
        "summary": {
            "buckets": summary, "overdue_count": overdue_count,
            "overdue_amt": round(overdue_amt, 2), "stale_count": stale_count,
        },
        "rows": page_rows, "total": total, "page": page, "size": size,
    }


# ---------------- 合同冗余判定 ----------------
GRADE_FACTOR = {"S": 1.5, "A": 1.0, "B": 0.8, "C": 0.5}


def grade_of(fc):
    if fc is None:
        return None
    if fc > 500:
        return "S"
    if fc > 200:
        return "A"
    if fc > 50:
        return "B"
    return "C"


def factor_of(grade):
    return GRADE_FACTOR.get(grade, 0.5)


def compute_redundancy(base):
    c = conn_db()
    rows = c.execute("""SELECT product_code, operator, operator_group, factory,
        outstanding_qty, unit_price, delivery_date FROM raw_orders""").fetchall()
    fc_rows = c.execute("SELECT product_code, operator, forecast_qty FROM sales_forecast").fetchall()
    c.close()
    try:
        base_d = datetime.date.fromisoformat(base)
    except Exception:
        base_d = datetime.date.today()

    fcmap = {}      # (product, operator) -> qty
    fctotal = {}    # product -> 总预测
    for r in fc_rows:
        pc, op, q = r["product_code"], r["operator"], r["forecast_qty"]
        if q is None:
            continue
        fcmap[(pc, op)] = (fcmap.get((pc, op), 0) or 0) + q
        fctotal[pc] = (fctotal.get(pc, 0) or 0) + q

    # 公司级冗余（按产品汇总未到货量/额）
    prod_qty = {}
    prod_amt = {}
    by_group = {}
    by_factory = {}
    by_sku_op = {}
    cycle_summary = {k: {"count": 0, "amt": 0.0} for k in
                     ["未到期", "逾期1-90", "91-180", "181-270", "271-365", "365以上"]}

    for r in rows:
        pc, op, grp, fac = r["product_code"], r["operator"], r["operator_group"], r["factory"]
        oq = r["outstanding_qty"] or 0
        if oq <= 0:
            continue
        price = r["unit_price"] or 0
        amt = oq * price
        prod_qty[pc] = prod_qty.get(pc, 0) + oq
        prod_amt[pc] = prod_amt.get(pc, 0) + amt
        g = by_group.setdefault(grp or "未填", {"qty": 0, "amt": 0, "count": 0})
        g["qty"] += oq; g["amt"] += amt; g["count"] += 1
        f = by_factory.setdefault(fac or "未填", {"qty": 0, "amt": 0, "count": 0})
        f["qty"] += oq; f["amt"] += amt; f["count"] += 1
        sk = by_sku_op.setdefault((pc, op or "未填"), {"qty": 0, "amt": 0, "count": 0})
        sk["qty"] += oq; sk["amt"] += amt; sk["count"] += 1
        # 交付周期分档
        dd = r["delivery_date"]
        if dd:
            try:
                dd_d = datetime.date.fromisoformat(parse_date(dd))
                cyc = (base_d - dd_d).days
                if cyc <= 0:
                    key = "未到期"
                elif cyc <= 90:
                    key = "逾期1-90"
                elif cyc <= 180:
                    key = "91-180"
                elif cyc <= 270:
                    key = "181-270"
                elif cyc <= 365:
                    key = "271-365"
                else:
                    key = "365以上"
                cycle_summary[key]["count"] += 1
                cycle_summary[key]["amt"] += amt
            except Exception:
                pass

    # 产品级接单判断
    def decision_for(pc):
        fc_total = fctotal.get(pc)
        if not fc_total:
            return "待填预测", None
        grade = grade_of(fc_total)
        safety = (fc_total / 3.0) * factor_of(grade)
        red = prod_qty.get(pc, 0)
        if red > safety:
            return "不可接单", grade
        return "可接单", grade

    by_product = []
    for pc, qty in prod_qty.items():
        dec, grade = decision_for(pc)
        by_product.append({
            "product_code": pc, "qty": round(qty, 0), "amt": round(prod_amt[pc], 2),
            "forecast": round(fctotal.get(pc, 0) or 0, 0), "grade": grade or "",
            "decision": dec,
        })
    by_product.sort(key=lambda x: -x["amt"])

    by_sku_op_list = []
    for (pc, op), v in by_sku_op.items():
        dec, grade = decision_for(pc)
        by_sku_op_list.append({
            "product_code": pc, "operator": op, "qty": round(v["qty"], 0),
            "amt": round(v["amt"], 2), "forecast": round(fcmap.get((pc, op), 0) or 0, 0),
            "grade": grade or "", "decision": dec,
        })
    by_sku_op_list.sort(key=lambda x: -x["amt"])

    by_group_list = [{"name": k, "qty": round(v["qty"], 0), "amt": round(v["amt"], 2),
                      "count": v["count"]} for k, v in by_group.items()]
    by_group_list.sort(key=lambda x: -x["amt"])
    by_factory_list = [{"name": k, "qty": round(v["qty"], 0), "amt": round(v["amt"], 2),
                        "count": v["count"]} for k, v in by_factory.items()]
    by_factory_list.sort(key=lambda x: -x["amt"])

    total_qty = sum(prod_qty.values())
    total_amt = sum(prod_amt.values())
    no_take = sum(1 for p in by_product if p["decision"] == "不可接单")
    return {
        "base": base,
        "summary": {
            "total_qty": round(total_qty, 0), "total_amt": round(total_amt, 2),
            "product_count": len(by_product), "no_take_count": no_take,
            "cycle": {k: {"count": v["count"], "amt": round(v["amt"], 2)} for k, v in cycle_summary.items()},
        },
        "by_group": by_group_list,
        "by_factory": by_factory_list,
        "by_product": by_product[:300],
        "by_sku_operator": by_sku_op_list[:300],
    }


# ---------------- 维度选项 ----------------
def distinct_values(col):
    c = conn_db()
    rows = c.execute("SELECT DISTINCT %s FROM raw_orders WHERE %s IS NOT NULL AND %s <> '' ORDER BY %s"
                     % (col, col, col, col)).fetchall()
    c.close()
    return [r[0] for r in rows]


# ---------------- 鉴权 ----------------
def get_session(handler):
    cookie = handler.headers.get("Cookie", "")
    m = re.search(r"sid=([A-Za-z0-9]+)", cookie)
    if not m:
        return None
    token = m.group(1)
    s = SESSIONS.get(token)
    if not s:
        return None
    if s["exp"] < time.time():
        SESSIONS.pop(token, None)
        return None
    return s


def set_session_cookie(handler, username, role, display_name):
    token = secrets.token_hex(16)
    SESSIONS[token] = {"username": username, "role": role,
                       "display_name": display_name, "exp": time.time() + SESSION_TTL}
    handler.send_header("Set-Cookie", "sid=%s; Path=/; Max-Age=%d; HttpOnly" % (token, SESSION_TTL))


def require_login(handler):
    if not AUTH_ENABLED:
        # 免登录模式：默认以管理员身份访问（无需账号密码）
        return {"username": "local", "role": "admin",
                "display_name": "本地用户", "exp": time.time() + 999999999}
    s = get_session(handler)
    if not s:
        send_json(handler, {"error": "未登录"}, 401)
        return None
    return s


def can_write(handler, s, path):
    if not AUTH_ENABLED:
        return True
    role = s["role"]
    if role == "admin":
        return True
    if role == "跟单员":
        # 跟单员可录入/编辑 原始数据 与 发货需求；不可删除、不可改产品/预测
        if path.startswith("/api/raw") and handler.command in ("POST", "PUT"):
            return True
        if path.startswith("/api/shipping") and handler.command == "POST":
            return True
        if path == "/api/proc/stage":
            return True
        return False
    return False  # viewer 只读


# ---------------- 响应 ----------------
def send_json(handler, obj, code=200):
    body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
    handler.send_response(code)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def read_body(handler):
    length = int(handler.headers.get("Content-Length", 0) or 0)
    if length == 0:
        return {}
    raw = handler.rfile.read(length).decode("utf-8")
    try:
        return json.loads(raw)
    except Exception:
        return {}


def read_body_bytes(handler):
    length = int(handler.headers.get("Content-Length", 0) or 0)
    if length == 0:
        return b""
    return handler.rfile.read(length)


def parse_multipart(data, boundary):
    """解析 multipart/form-data，返回 {field_name: (filename, bytes)}。
    Python 3.13 已移除 cgi 模块，故手写。"""
    parts = {}
    delim = b"--" + boundary
    for seg in data.split(delim):
        if not seg.startswith(b"\r\n"):
            continue
        hidx = seg.find(b"\r\n\r\n")
        if hidx == -1:
            continue
        header = seg[2:hidx].decode("utf-8", "ignore")
        body = seg[hidx + 4:]
        if body.endswith(b"\r\n"):
            body = body[:-2]
        m = re.search(r'name="([^"]*)"', header)
        if not m:
            continue
        name = m.group(1)
        fm = re.search(r'filename="([^"]*)"', header)
        parts[name] = (fm.group(1) if fm else None, body)
    return parts


# Excel 导入：列名别名（同时兼容网页中文表头与真实工作簿表头）
RAW_ALIASES = {
    "contract_no": ["合同号", "采购合同"],
    "product_code": ["产品编号"],
    "team": ["团队"],
    "operator": ["运营专员", "运营"],
    "operator_group": ["运营组别", "跟单组别"],
    "factory": ["厂商简称", "厂商"],
    "order_qty": ["订单总数量", "合同数量"],
    "delivered_qty": ["已交付数量"],
    "undelivered_qty": ["未入库数量"],
    "outstanding_qty": ["未到货总数量"],
    "unit_price": ["采购单价"],
    "contract_date": ["合同日期"],
    "delivery_date": ["交货日期"],
    "customs_material": ["报关材质"],
    "followup_date": ["最新跟进日期"],
    "followup_conclusion": ["跟进结论"],
    "followup_note": ["跟进备注"],
    "is_new": ["是否新单"],
}


def import_rows_generic(data, aliases):
    """解析 xlsx 字节为 [(行号, {字段:值}), ...]，按 aliases 中文表头映射列。第一行为表头。"""
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        return []
    keymap = {}  # 列索引 -> 字段
    for ci, h in enumerate(header):
        hs = str(h).strip() if h is not None else ""
        if not hs:
            continue
        for key, aliases_list in aliases.items():
            names = [key] + list(aliases_list)
            # 子串匹配：兼容「产品编号(SKU)*」这类带后缀/星号的表头
            if any(a and (a == hs or a in hs or hs in a) for a in names):
                keymap[ci] = key
                break
    out = []
    for ri, row in enumerate(rows_iter, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        obj = {}
        for ci, key in keymap.items():
            val = row[ci] if ci < len(row) else None
            if val is None:
                val = ""
            elif isinstance(val, (datetime.datetime, datetime.date)):
                val = val.strftime("%Y-%m-%d")
            else:
                val = str(val).strip()
            obj[key] = val
        out.append((ri, obj))
    return out


def import_excel_rows(data):
    """解析原始数据 xlsx（兼容历史调用）。"""
    return import_rows_generic(data, RAW_ALIASES)


def insert_generic(c, tname, cols, d):
    """通用插入：仅写非空列。"""
    cols2 = [col for col in cols if col in d and d[col] is not None]
    vals = [d.get(col) for col in cols2]
    ph = ",".join(["?"] * len(cols2))
    c.execute("INSERT INTO %s (%s) VALUES (%s)" % (tname, ",".join(cols2), ph), vals)



def insert_raw(c, d):
    cols = [col for col in RAW_COLS if col in d and d[col] is not None]
    vals = [d.get(col) for col in cols]
    ph = ",".join(["?"] * len(cols))
    cur = c.execute("INSERT INTO raw_orders (%s) VALUES (%s)" % (",".join(cols), ph), vals)
    return cur.lastrowid


# ---------------- 派生计算（02/03/04/06/10） ----------------
def load_products_map():
    c = conn_db()
    rows = c.execute("SELECT product_code, craft_category FROM products").fetchall()
    c.close()
    return {r["product_code"]: (r["craft_category"] or "") for r in rows}


def contract_total_map():
    c = conn_db()
    rows = c.execute("SELECT contract_no, COALESCE(SUM(order_qty),0) AS t FROM raw_orders GROUP BY contract_no").fetchall()
    c.close()
    return {r["contract_no"]: r["t"] for r in rows}


def pkey_of(product_code, contract_no):
    return (product_code or "") + "||" + (contract_no or "")


def load_proc_actual():
    c = conn_db()
    rows = c.execute("SELECT pkey, stage_idx, actual_date FROM proc_stages").fetchall()
    c.close()
    d = {}
    for r in rows:
        if r["actual_date"]:
            d.setdefault(r["pkey"], {})[r["stage_idx"]] = r["actual_date"]
    return d


def compute_order_detail(rows, base=None):
    pmap = load_products_map()
    ctmap = contract_total_map()
    out = []
    for r in rows:
        cat = pmap.get(r.get("product_code"), "")
        ct = ctmap.get(r.get("contract_no"), 0)
        tier = PL.resolve_tier(ct)
        oq = float(r.get("order_qty") or 0)
        dq = float(r.get("delivered_qty") or 0)
        cr = (dq / oq) if oq else 0
        f = PL.order_detail_fields(r.get("order_qty"), r.get("delivered_qty"),
                                   r.get("outstanding_qty"), r.get("delivery_date"),
                                   cr, ct, tier, r.get("is_new"), None, base)
        row = dict(r)
        row["craft_category"] = cat
        row["contract_total"] = ct
        row["tier"] = tier
        row.update(f)
        out.append(row)
    return out


def compute_proc(rows, base=None):
    pmap = load_products_map()
    ctmap = contract_total_map()
    actual = load_proc_actual()
    out = []
    for r in rows:
        rid = r.get("id")
        cat = pmap.get(r.get("product_code"), "")
        tier = PL.resolve_tier(ctmap.get(r.get("contract_no"), 0))
        pr = PL.compute_proc(cat, tier, r.get("is_new") or "否",
                             r.get("contract_date"), r.get("delivery_date"),
                             actual.get(rid, {}), base)
        out.append({
            "id": rid, "contract_no": r.get("contract_no"),
            "product_code": r.get("product_code"), "factory": r.get("factory"),
            "craft_category": cat, "tier": tier,
            "delivery_date": r.get("delivery_date"), "is_new": r.get("is_new"),
            "outstanding_qty": r.get("outstanding_qty"), **pr,
        })
    return out


def compute_gantt(rows, base=None, n=22):
    pmap = load_products_map()
    ctmap = contract_total_map()
    actual = load_proc_actual()
    weeks = PL.gantt_weeks(base, n)
    out = []
    for r in rows:
        rid = pkey_of(r.get("product_code"), r.get("contract_no"))
        cat = pmap.get(r.get("product_code"), "")
        tier = PL.resolve_tier(ctmap.get(r.get("contract_no"), 0))
        pr = PL.compute_proc(cat, tier, r.get("is_new") or "否",
                             r.get("contract_date"), r.get("delivery_date"),
                             actual.get(rid, {}), base)
        plan_starts = [s["plan_start"] for s in pr["stages"] if s["applicable"]]
        cells = [PL.gantt_cell(plan_starts, w["week_end"]) for w in weeks]
        out.append({
            "id": rid, "contract_no": r.get("contract_no"),
            "product_code": r.get("product_code"), "factory": r.get("factory"),
            "craft_category": cat, "delivery_date": r.get("delivery_date"),
            "outstanding_qty": r.get("outstanding_qty"),
            "block": pr["block"], "progress": pr["progress"],
            "plan_starts": plan_starts, "cells": cells,
        })
    return {"weeks": weeks, "rows": out}


def load_products_full():
    c = conn_db()
    rows = c.execute("SELECT product_code, product_name, craft_category, spu FROM products").fetchall()
    c.close()
    return {r["product_code"]: {
        "product_name": r["product_name"] or "",
        "craft_category": r["craft_category"] or "",
        "spu": r["spu"] or "",
    } for r in rows}


def aggregate_proc(search="", filters=None):
    """工序进度模块独立聚合：按 (产品编号, 合同号) 汇总 proc_detail，
    关联产品资料(取 SPU/工艺品类)、proc_meta(手填字段)、proc_stages(手动工序进度)。
    返回按 合同号/产品编号 排序的分组列表（含分页前全量，供端点做筛选与分页）。"""
    filters = filters or {}
    c = conn_db()
    sql = "SELECT * FROM proc_detail"
    args = []
    where = []
    if search:
        like = "%" + search + "%"
        where.append("(contract_no LIKE ? OR product_code LIKE ? OR spu LIKE ? OR craft_category LIKE ?)")
        args += [like, like, like, like]
    if where:
        sql += " WHERE " + " AND ".join(where)
    details = [dict(r) for r in c.execute(sql, args).fetchall()]

    groups = {}
    for d in details:
        pk = pkey_of(d.get("product_code"), d.get("contract_no"))
        g = groups.setdefault(pk, {
            "pkey": pk, "contract_no": d.get("contract_no") or "",
            "product_code": d.get("product_code") or "",
            "spu_list": [], "cat_list": [], "order_qty": 0.0,
            "contract_dates": [], "ship_dates": [], "detail_ids": [], "detail_count": 0,
        })
        try:
            g["order_qty"] += float(d.get("order_qty") or 0)
        except Exception:
            pass
        if d.get("spu"):
            g["spu_list"].append(d["spu"])
        if d.get("craft_category"):
            g["cat_list"].append(d["craft_category"])
        if d.get("contract_date"):
            g["contract_dates"].append(d["contract_date"])
        if d.get("ship_date"):
            g["ship_dates"].append(d["ship_date"])
        g["detail_ids"].append(d["id"])
        g["detail_count"] += 1

    pmap = load_products_full()
    meta_rows = c.execute("SELECT * FROM proc_meta").fetchall()
    meta = {(r["product_code"], r["contract_no"]): dict(r) for r in meta_rows}
    actual = load_proc_actual()
    c.close()

    out = []
    for pk, g in groups.items():
        pc = g["product_code"]; cn = g["contract_no"]
        pinfo = pmap.get(pc, {})
        spu = next((s for s in g["spu_list"] if s), "") or pinfo.get("spu") or pinfo.get("product_name") or ""
        cat = next((s for s in g["cat_list"] if s), "") or pinfo.get("craft_category") or ""
        contract_date = min(g["contract_dates"]) if g["contract_dates"] else ""
        ship_date = min(g["ship_dates"]) if g["ship_dates"] else ""
        tier = PL.resolve_tier(g["order_qty"])
        pr = PL.compute_proc(cat, tier, "否", contract_date, ship_date, actual.get(pk, {}), None)
        m = meta.get((pc, cn), {})
        out.append({
            "pkey": pk, "contract_no": cn, "product_code": pc, "spu": spu,
            "craft_category": cat, "tier": tier,
            "order_qty": round(g["order_qty"], 2),
            "contract_date": contract_date, "ship_date": ship_date,
            "customer_complaint": m.get("customer_complaint") or "",
            "inspection_focus": m.get("inspection_focus") or "",
            "inspection_time": m.get("inspection_time") or "",
            "inspection_result": m.get("inspection_result") or "",
            "detail_count": g["detail_count"], "detail_ids": g["detail_ids"],
            **pr,
        })
    # 列名筛选（应用层，针对聚合后的字段）
    for col in ("contract_no", "product_code", "spu", "craft_category", "inspection_focus"):
        fv = (filters.get(col) or "").strip()
        if fv:
            out = [r for r in out if fv.lower() in (str(r.get(col) or "")).lower()]
    out.sort(key=lambda r: (r["contract_no"], r["product_code"]))
    return out


def compute_contract_summary(base=None):
    c = conn_db()
    rows = c.execute("""SELECT contract_no, factory,
        COUNT(*) AS lines, COUNT(DISTINCT product_code) AS skus,
        COALESCE(SUM(order_qty),0) AS order_qty,
        COALESCE(SUM(delivered_qty),0) AS delivered_qty,
        COALESCE(SUM(outstanding_qty),0) AS outstanding_qty,
        MIN(delivery_date) AS min_dd, MAX(delivery_date) AS max_dd
        FROM raw_orders GROUP BY contract_no ORDER BY outstanding_qty DESC""").fetchall()
    c.close()
    base_d = PL.parse_date(base or PL.today_str())
    out = []
    for r in rows:
        outstanding = float(r["outstanding_qty"] or 0)
        dd = PL.parse_date(r["max_dd"])
        days = (dd - base_d).days if (dd and base_d) else None
        if outstanding <= 0:
            cstatus = "已交清"
        elif days is None:
            cstatus = "未知"
        elif days < 0:
            cstatus = "已逾期"
        elif days <= PL.THRESHOLDS["urgent"]:
            cstatus = "紧急"
        elif days <= PL.THRESHOLDS["warn"]:
            cstatus = "预警"
        else:
            cstatus = "正常"
        oq = float(r["order_qty"] or 0)
        cr = (float(r["delivered_qty"] or 0) / oq) if oq else 0
        whole = "✔ 可整单出货" if outstanding <= 0 else ("缺 %.0f 件" % outstanding)
        out.append({
            "contract_no": r["contract_no"], "factory": r["factory"],
            "lines": r["lines"], "skus": r["skus"],
            "order_qty": r["order_qty"], "delivered_qty": r["delivered_qty"],
            "outstanding_qty": outstanding, "complete_rate": cr,
            "min_dd": r["min_dd"], "max_dd": r["max_dd"],
            "days_to_ship": days, "status": cstatus, "whole": whole,
        })
    return out


def compute_shipchk(start, end):
    c = conn_db()
    eff_start = start or "2000-01-01"
    eff_end = end or "2999-12-31"
    demand = c.execute("""SELECT sku AS product_code, team, COALESCE(SUM(ship_qty),0) AS qty
        FROM shipping_demands WHERE ship_date >= ? AND ship_date <= ?
        GROUP BY sku, team""", (eff_start, eff_end)).fetchall()
    supply = c.execute("""SELECT product_code, team,
        COALESCE(SUM(undelivered_qty),0) AS qty,
        COALESCE(SUM(outstanding_qty),0) AS out_qty,
        COALESCE(SUM(order_qty),0) AS ord_qty
        FROM raw_orders GROUP BY product_code, team""").fetchall()
    c.close()
    keys = {}
    for d in demand:
        k = (d["product_code"], d["team"])
        keys.setdefault(k, {})["demand"] = float(d["qty"])
    for s in supply:
        k = (s["product_code"], s["team"])
        keys.setdefault(k, {})
        keys[k]["supply"] = float(s["qty"])
        keys[k]["contract"] = float(s["ord_qty"])
        keys[k]["outstanding"] = float(s["out_qty"])
    rows = []
    demand_total = supply_total = 0.0
    gap_rows = 0
    for k, v in keys.items():
        if v.get("demand", 0) <= 0:
            continue
        dem = v["demand"]
        sup = v.get("supply", 0)
        diff = sup - dem
        demand_total += dem
        supply_total += sup
        status = "满足✔" if diff >= 0 else ("缺口 %.0f" % -diff)
        if diff < 0:
            gap_rows += 1
        rows.append({
            "product_code": k[0], "team": k[1],
            "demand": dem, "supply": sup, "contract": v.get("contract", 0),
            "outstanding": v.get("outstanding", 0), "diff": diff, "status": status,
        })
    rows.sort(key=lambda x: x["diff"])
    satisfy_rate = (len(rows) - gap_rows) / len(rows) if rows else 0
    return {
        "start": start, "end": end, "eff_start": eff_start, "eff_end": eff_end,
        "kpi": {
            "demand_total": demand_total, "supply_total": supply_total,
            "gap": supply_total - demand_total, "gap_rows": gap_rows,
            "rows": len(rows), "satisfy_rate": satisfy_rate,
        },
        "rows": rows,
    }


# ---------------- Handler ----------------
class Handler(BaseHTTPRequestHandler):
    def log_message(self, *a):
        pass

    def _serve_static(self, path):
        rel = path[len("/static/"):]
        fpath = os.path.join(STATIC, rel)
        fpath = os.path.normpath(fpath)
        if not fpath.startswith(STATIC) or not os.path.isfile(fpath):
            self.send_error(404)
            return
        ctype = "text/html; charset=utf-8"
        if fpath.endswith(".js"):
            ctype = "application/javascript; charset=utf-8"
        elif fpath.endswith(".css"):
            ctype = "text/css; charset=utf-8"
        with open(fpath, "rb") as f:
            data = f.read()
        self.send_response(200)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate")
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        u = urlparse(self.path)
        p = u.path
        if p in ("/", "/index.html"):
            with open(os.path.join(STATIC, "index.html"), "rb") as f:
                data = f.read()
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(data)))
            self.send_header("Cache-Control", "no-store, no-cache, must-revalidate")
            self.end_headers()
            self.wfile.write(data)
            return
        if p.startswith("/static/"):
            self._serve_static(p)
            return
        # 以下接口需要登录
        s = require_login(self)
        if not s:
            return
        q = parse_qs(u.query)
        if p == "/api/me":
            send_json(self, {"username": s["username"], "role": s["role"], "display_name": s["display_name"]})
            return
        if p == "/api/kpi":
            send_json(self, compute_kpi())
            return
        if p == "/api/columns":
            send_json(self, {"cols": RAW_COLS, "labels": RAW_LABELS})
            return
        if p == "/api/filters":
            send_json(self, {
                "operator": distinct_values("operator"),
                "operator_group": distinct_values("operator_group"),
                "factory": distinct_values("factory"),
            })
            return
        if p == "/api/raw":
            page = int(q.get("page", ["1"])[0])
            size = int(q.get("size", ["100"])[0])
            search = q.get("search", [""])[0].strip()
            flt = {col: q.get(col, [""])[0].strip() for col in FILTER_COLS}
            flt = {k: v for k, v in flt.items() if v}
            where, args = build_raw_where(search, flt)
            c = conn_db()
            total = c.execute("SELECT COUNT(*) AS n FROM raw_orders " + where, args).fetchone()["n"]
            rows = c.execute(
                "SELECT * FROM raw_orders " + where +
                " ORDER BY id DESC LIMIT ? OFFSET ?",
                args + [size, (page - 1) * size]).fetchall()
            c.close()
            send_json(self, {"rows": [dict(r) for r in rows],
                             "total": total, "page": page, "size": size,
                             "labels": RAW_LABELS, "cols": RAW_COLS})
            return
        if p == "/api/products":
            c = conn_db()
            rows = c.execute("SELECT * FROM products ORDER BY id DESC LIMIT 500").fetchall()
            c.close()
            send_json(self, {"rows": [dict(r) for r in rows]})
            return
        if p == "/api/shipping":
            c = conn_db()
            rows = c.execute("SELECT * FROM shipping_demands ORDER BY id DESC LIMIT 500").fetchall()
            c.close()
            send_json(self, {"rows": [dict(r) for r in rows]})
            return
        if p == "/api/forecast":
            c = conn_db()
            rows = c.execute("SELECT * FROM sales_forecast ORDER BY id DESC LIMIT 500").fetchall()
            c.close()
            send_json(self, {"rows": [dict(r) for r in rows]})
            return
        # 三张表的 Excel 导入模板下载（含中文表头，* 表示必填）
        if p in ("/api/products/template", "/api/shipping/template", "/api/forecast/template"):
            key = p.split("/")[-2]
            cfg = IMPORT_CFG[key]
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([cfg["labels"][c] for c in cfg["cols"]])
            # 加一行示例，便于理解填写方式
            example = {
                "products": ["NBCH028BREOL-V1", "编藤单椅", "编藤"],
                "shipping": ["NBCH028BREOL-V1", "户外组", "张三", "佛山厂", "李四", "120", "2026-09-01"],
                "forecast": ["NBCH028BREOL-V1", "张三", "300"],
            }[key]
            ws.append(example)
            ws.freeze_panes = "A2"
            bio = io.BytesIO()
            wb.save(bio)
            data = bio.getvalue()
            fname = {"products": "产品资料表_导入模板", "shipping": "发货需求表_导入模板",
                     "forecast": "销售预测表_导入模板"}[key]
            fname_en = {"products": "products_template", "shipping": "shipping_template",
                        "forecast": "forecast_template"}[key]
            disp = "attachment; filename=%s.xlsx; filename*=UTF-8''%s" % (fname_en, quote(fname + ".xlsx"))
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", disp)
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if p == "/api/overdue":
            base = q.get("base", [""])[0].strip() or today_str()
            bucket = q.get("bucket", [""])[0].strip()
            group = q.get("group", [""])[0].strip()
            operator = q.get("operator", [""])[0].strip()
            factory = q.get("factory", [""])[0].strip()
            page = int(q.get("page", ["1"])[0])
            size = int(q.get("size", ["200"])[0])
            send_json(self, compute_overdue(base, group, operator, factory, bucket, page, size))
            return
        if p == "/api/redundancy":
            base = q.get("base", [""])[0].strip() or today_str()
            send_json(self, compute_redundancy(base))
            return
        if p == "/api/order-detail":
            page = int(q.get("page", ["1"])[0]); size = int(q.get("size", ["100"])[0])
            search = q.get("search", [""])[0].strip()
            flt = {col: q.get(col, [""])[0].strip() for col in FILTER_COLS}
            flt = {k: v for k, v in flt.items() if v}
            where, args = build_raw_where(search, flt)
            c = conn_db()
            total = c.execute("SELECT COUNT(*) AS n FROM raw_orders " + where, args).fetchone()["n"]
            rows = c.execute("SELECT * FROM raw_orders " + where +
                             " ORDER BY id DESC LIMIT ? OFFSET ?", args + [size, (page - 1) * size]).fetchall()
            c.close()
            data = compute_order_detail([dict(r) for r in rows], today_str())
            send_json(self, {"rows": data, "total": total, "page": page, "size": size,
                             "labels": RAW_LABELS, "cols": RAW_COLS})
            return
        if p == "/api/proc":
            page = int(q.get("page", ["1"])[0]); size = int(q.get("size", ["100"])[0])
            search = q.get("search", [""])[0].strip()
            filters = {
                "contract_no": q.get("fc", [""])[0].strip(),
                "product_code": q.get("fp", [""])[0].strip(),
                "spu": q.get("fs", [""])[0].strip(),
                "craft_category": q.get("fcat", [""])[0].strip(),
                "inspection_focus": q.get("finsp", [""])[0].strip(),
            }
            filters = {k: v for k, v in filters.items() if v}
            groups = aggregate_proc(search, filters)
            total = len(groups)
            all_keys = [g["pkey"] for g in groups]
            page_groups = groups[(page - 1) * size: page * size]
            send_json(self, {"rows": page_groups, "total": total, "page": page, "size": size,
                             "all_keys": all_keys,
                             "stages": PL.STAGES, "stage_names_by_cat": PL.CAT_STAGE_NAME,
                             "inspection_focus_options": INSPECTION_FOCUS})
            return
        if p == "/api/proc/detail":
            pc = q.get("product_code", [""])[0].strip()
            cn = q.get("contract_no", [""])[0].strip()
            c = conn_db()
            rows = c.execute(
                "SELECT id, contract_no, product_code, spu, order_qty, contract_date, ship_date, craft_category, created_at "
                "FROM proc_detail WHERE product_code=? AND contract_no=? ORDER BY id",
                (pc, cn)).fetchall()
            c.close()
            send_json(self, {"rows": [dict(r) for r in rows],
                             "product_code": pc, "contract_no": cn})
            return
        if p == "/api/proc/template":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([PROC_LABELS[c] for c in PROC_COLS])
            ws.append(["PO241218001", "NBCH028BREOL-V1", "SPU-A01", "120", "2026-01-10", "2026-09-01", "编藤"])
            ws.freeze_panes = "A2"
            bio = io.BytesIO()
            wb.save(bio)
            data = bio.getvalue()
            fname = "工序进度明细_导入模板"
            fname_en = "proc_detail_template"
            disp = "attachment; filename=%s.xlsx; filename*=UTF-8''%s" % (fname_en, quote(fname + ".xlsx"))
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", disp)
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if p == "/api/gantt":
            base = q.get("base", [""])[0].strip() or today_str()
            page = int(q.get("page", ["1"])[0]); size = int(q.get("size", ["200"])[0])
            search = q.get("search", [""])[0].strip()
            flt = {col: q.get(col, [""])[0].strip() for col in FILTER_COLS}
            flt = {k: v for k, v in flt.items() if v}
            where, args = build_raw_where(search, flt)
            c = conn_db()
            rows = c.execute("SELECT * FROM raw_orders " + where +
                             " ORDER BY id DESC LIMIT ? OFFSET ?", args + [size, (page - 1) * size]).fetchall()
            c.close()
            send_json(self, compute_gantt([dict(r) for r in rows], base, 22))
            return
        if p == "/api/contract":
            base = q.get("base", [""])[0].strip() or today_str()
            send_json(self, compute_contract_summary(base))
            return
        if p == "/api/shipchk":
            start = q.get("start", [""])[0].strip()
            end = q.get("end", [""])[0].strip()
            send_json(self, compute_shipchk(start, end))
            return
        if p == "/api/params":
            send_json(self, {
                "stages": PL.STAGES,
                "cat_stage_name": PL.CAT_STAGE_NAME,
                "cat_order": PL.CAT_ORDER,
                "tiers": PL.TIERS, "news": PL.NEWS,
                "thresholds": PL.THRESHOLDS,
                "tact": [{"cat": k[0], "tier": k[1], "is_new": k[2], "offsets": v} for k, v in PL.TACT.items()],
                "default_lead_days": PL.DEFAULT_LEAD_DAYS,
            })
            return
        self.send_error(404)

    def do_POST(self):
        u = urlparse(self.path)
        p = u.path
        if p == "/api/login":
            # 免登录模式：不校验密码，直接以管理员身份进入（即使页面仍显示登录框，点登录也能进）
            if not AUTH_ENABLED:
                self.send_response(200)
                set_session_cookie(self, "local", "admin", "本地用户")
                send_json(self, {"ok": True, "role": "admin", "username": "local",
                                 "display_name": "本地用户"})
                return
            d = read_body(self)
            c = conn_db()
            row = c.execute("SELECT * FROM users WHERE username=?", (d.get("username", ""),)).fetchone()
            c.close()
            if not row or not check_pw(d.get("password", ""), row["password_hash"]):
                send_json(self, {"error": "用户名或密码错误"}, 401)
                return
            self.send_response(200)
            set_session_cookie(self, row["username"], row["role"], row["display_name"])
            send_json(self, {"ok": True, "role": row["role"], "username": row["username"],
                             "display_name": row["display_name"]})
            return
        if p == "/api/logout":
            cookie = self.headers.get("Cookie", "")
            m = re.search(r"sid=([A-Za-z0-9]+)", cookie)
            if m:
                SESSIONS.pop(m.group(1), None)
            send_json(self, {"ok": True})
            return
        if p == "/api/backup":
            s = require_login(self)
            if not s:
                return
            if not can_write(self, s, p):
                send_json(self, {"error": "无权限"}, 403)
                return
            dst = backup_db()
            send_json(self, {"ok": True, "path": dst})
            return
        # 其余写操作需登录+权限
        s = require_login(self)
        if not s:
            return
        if not can_write(self, s, p):
            send_json(self, {"error": "无权限（当前角色不可写）"}, 403)
            return
        if p == "/api/raw":
            d = read_body(self)
            errs = validate_raw(d)
            if errs:
                send_json(self, {"error": "校验失败", "details": errs}, 400)
                return
            cd = clean_raw(d)
            c = conn_db()
            rid = insert_raw(c, cd)
            c.commit(); c.close()
            send_json(self, {"ok": True, "id": rid})
            return
        if p == "/api/raw/bulk":
            d = read_body(self)
            rows_in = d.get("rows", []) if isinstance(d, dict) else []
            ok_rows, errs = [], []
            c = conn_db()
            for i, r in enumerate(rows_in, 1):
                e = validate_raw(r)
                if e:
                    errs.append({"row": i, "errors": e})
                else:
                    cd = clean_raw(r)
                    insert_raw(c, cd)
                    ok_rows.append(i)
            c.commit(); c.close()
            send_json(self, {"ok": len(errs) == 0, "inserted": len(ok_rows),
                             "errors": errs, "error_count": len(errs)})
            return
        if p == "/api/raw/import":
            ctype = self.headers.get("Content-Type", "")
            m = re.search(r"boundary=([^;]+)", ctype)
            if not m:
                send_json(self, {"error": "请使用 multipart/form-data 上传文件"}, 400)
                return
            data = read_body_bytes(self)
            parts = parse_multipart(data, m.group(1).strip().encode())
            fpart = parts.get("file")
            if not fpart or not fpart[1]:
                send_json(self, {"error": "未收到文件"}, 400)
                return
            try:
                pairs = import_excel_rows(fpart[1])
            except Exception as e:
                send_json(self, {"error": "解析 Excel 失败：" + str(e)}, 400)
                return
            if not pairs:
                send_json(self, {"error": "文件中未解析到数据行（请确认第一行为表头）"}, 400)
                return
            ok_rows, errs = [], []
            c = conn_db()
            for ri, obj in pairs:
                e = validate_raw(obj)
                if e:
                    errs.append({"row": ri, "errors": e})
                else:
                    insert_raw(c, clean_raw(obj))
                    ok_rows.append(ri)
            c.commit(); c.close()
            send_json(self, {"ok": len(errs) == 0, "inserted": len(ok_rows),
                             "errors": errs, "error_count": len(errs)})
            return
        # 三张表的 Excel 批量导入
        if p in ("/api/products/import", "/api/shipping/import", "/api/forecast/import"):
            key = p.split("/")[-2]
            cfg = IMPORT_CFG[key]
            m = re.search(r"boundary=([^;]+)", self.headers.get("Content-Type", ""))
            if not m:
                send_json(self, {"error": "请使用 multipart/form-data 上传文件"}, 400)
                return
            data = read_body_bytes(self)
            parts = parse_multipart(data, m.group(1).strip().encode())
            fpart = parts.get("file")
            if not fpart or not fpart[1]:
                send_json(self, {"error": "未收到文件"}, 400)
                return
            try:
                pairs = import_rows_generic(fpart[1], cfg["aliases"])
            except Exception as e:
                send_json(self, {"error": "解析 Excel 失败：" + str(e)}, 400)
                return
            if not pairs:
                send_json(self, {"error": "文件中未解析到数据行（请确认第一行为中文表头，如：产品编号、中文品名、工艺品类）"}, 400)
                return
            ok_rows, errs = [], []
            # 字段名 -> 中文表头（去掉 *），用于把校验报错翻译得可读
            zh = {col: lab.replace("*", "") for col, lab in cfg["labels"].items()}
            def tr(err):
                for col, lab in zh.items():
                    if err.startswith(col + " "):
                        return lab + err[len(col):]
                return err
            c = conn_db()
            for ri, obj in pairs:
                e = validate_generic(obj, cfg["cols"], cfg["required"])
                if e:
                    errs.append({"row": ri, "errors": [tr(x) for x in e]})
                    continue
                try:
                    insert_generic(c, cfg["table"], cfg["cols"], obj)
                    ok_rows.append(ri)
                except Exception as ex:
                    errs.append({"row": ri, "errors": [str(ex)]})
            c.commit(); c.close()
            send_json(self, {"ok": len(errs) == 0, "inserted": len(ok_rows),
                             "errors": errs, "error_count": len(errs)})
            return
        if p in ("/api/products", "/api/shipping", "/api/forecast"):
            tname = {"/api/products": "products", "/api/shipping": "shipping_demands",
                     "/api/forecast": "sales_forecast"}[p]
            cols_map = {"products": PROD_COLS, "shipping_demands": SHIP_COLS, "sales_forecast": FC_COLS}
            d = read_body(self)
            cd = {col: d.get(col) for col in cols_map[tname] if col in d}
            errs = validate_generic(cd, cols_map[tname],
                                    [c for c in cols_map[tname] if c not in ("product_name", "craft_category", "team", "operator", "factory", "follower", "ship_date", "customs_material")])
            if errs:
                send_json(self, {"error": "校验失败", "details": errs}, 400)
                return
            cols = [c for c in cols_map[tname] if c in cd and cd[c] is not None]
            vals = [cd.get(c) for c in cols]
            ph = ",".join(["?"] * len(cols))
            c = conn_db()
            c.execute("INSERT INTO %s (%s) VALUES (%s)" % (tname, ",".join(cols), ph), vals)
            c.commit(); c.close()
            send_json(self, {"ok": True})
            return
        if p == "/api/proc/stage":
            d = read_body(self)
            pkey = to_str(d.get("pkey"))
            try:
                # 注意：0 是合法值，不能用 "x or -1" 把 0 替换成 -1
                sidx_raw = d.get("stage_idx", "")
                sidx = int(sidx_raw) if str(sidx_raw).strip() != "" else -1
            except Exception:
                send_json(self, {"error": "参数无效"}, 400)
                return
            actual = to_str(d.get("actual_date"))
            if not pkey or sidx < 0 or sidx > 8:
                send_json(self, {"error": "参数无效（pkey/stage_idx）"}, 400)
                return
            c = conn_db()
            if actual:
                c.execute("INSERT OR REPLACE INTO proc_stages (pkey, stage_idx, actual_date) VALUES (?,?,?)",
                          (pkey, sidx, actual))
            else:
                c.execute("DELETE FROM proc_stages WHERE pkey=? AND stage_idx=?", (pkey, sidx))
            c.commit(); c.close()
            send_json(self, {"ok": True})
            return
        if p == "/api/proc/stage-batch":
            d = read_body(self)
            if not isinstance(d, dict):
                d = {}
            try:
                # 注意：0 是合法值，不能用 "x or -1" 把 0 替换成 -1
                sidx_raw = d.get("stage_idx", "")
                sidx = int(sidx_raw) if str(sidx_raw).strip() != "" else -1
            except Exception:
                send_json(self, {"error": "stage_idx 参数无效"}, 400)
                return
            actual = to_str(d.get("actual_date"))
            if sidx < 0 or sidx > 8:
                send_json(self, {"error": "stage_idx 必须在 0..8"}, 400)
                return
            c = conn_db()
            ok = 0; fail = 0; err = ""
            pkeys = []
            if d.get("scope") == "all":
                # 把当前搜索/筛选条件下所有聚合分组的 pkey 取出来再批量更新
                search = (d.get("search") or "").strip()
                filters = {
                    "contract_no": (d.get("fc") or "").strip(),
                    "product_code": (d.get("fp") or "").strip(),
                    "spu": (d.get("fs") or "").strip(),
                    "craft_category": (d.get("fcat") or "").strip(),
                    "inspection_focus": (d.get("finsp") or "").strip(),
                }
                filters = {k: v for k, v in filters.items() if v}
                groups = aggregate_proc(search, filters)
                pkeys = [g["pkey"] for g in groups]
            else:
                pkeys = [str(x) for x in d.get("pkeys", []) if str(x).strip()]
            if not pkeys:
                c.close(); send_json(self, {"error": "未提供有效 pkeys（请先勾选行，或使用「当前搜索/筛选下的全部」）"}, 400); return
            try:
                if actual:
                    c.execute("BEGIN")
                    c.executemany(
                        "INSERT OR REPLACE INTO proc_stages (pkey, stage_idx, actual_date) VALUES (?,?,?)",
                        [(pk, sidx, actual) for pk in pkeys])
                    ok = len(pkeys)
                    c.execute("COMMIT")
                else:
                    # 留空 = 清除全部命中的工序
                    c.execute("BEGIN")
                    ph = ",".join("?" * len(pkeys))
                    c.execute("DELETE FROM proc_stages WHERE stage_idx=? AND pkey IN (%s)" % ph,
                              [sidx] + pkeys)
                    ok = c.execute("SELECT changes()").fetchone()[0]
                    c.execute("COMMIT")
            except Exception as e:
                c.execute("ROLLBACK")
                fail = len(pkeys); ok = 0
                err = str(e)
            c.close()
            send_json(self, {"ok": True, "updated": ok, "failed": fail,
                             "scope": "all" if d.get("scope") == "all" else "ids",
                             "target_count": len(pkeys),
                             "error": err if fail else ""})
            return
        if p == "/api/proc/import":
            m = re.search(r"boundary=([^;]+)", self.headers.get("Content-Type", ""))
            if not m:
                send_json(self, {"error": "请使用 multipart/form-data 上传文件"}, 400)
                return
            data = read_body_bytes(self)
            parts = parse_multipart(data, m.group(1).strip().encode())
            fpart = parts.get("file")
            if not fpart or not fpart[1]:
                send_json(self, {"error": "未收到文件"}, 400)
                return
            try:
                pairs = import_rows_generic(fpart[1], PROC_ALIASES)
            except Exception as e:
                send_json(self, {"error": "解析 Excel 失败：" + str(e)}, 400)
                return
            if not pairs:
                send_json(self, {"error": "文件中未解析到数据行（请确认第一行为中文表头，如：合同号、产品编号、订单数量）"}, 400)
                return
            ok_rows, errs = [], []
            batch = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
            c = conn_db()
            for ri, obj in pairs:
                e = validate_proc(obj)
                if e:
                    errs.append({"row": ri, "errors": e})
                    continue
                try:
                    cd = clean_proc(obj, batch)
                    insert_generic(c, "proc_detail", PROC_COLS + ["import_batch"], cd)
                    ok_rows.append(ri)
                except Exception as ex:
                    errs.append({"row": ri, "errors": [str(ex)]})
            c.commit(); c.close()
            send_json(self, {"ok": len(errs) == 0, "inserted": len(ok_rows),
                             "errors": errs, "error_count": len(errs)})
            return
        if p == "/api/proc/meta":
            d = read_body(self)
            pc = to_str(d.get("product_code"))
            cn = to_str(d.get("contract_no"))
            if not pc or not cn:
                send_json(self, {"error": "product_code/contract_no 必填"}, 400)
                return
            fields = {
                "customer_complaint": to_str(d.get("customer_complaint")),
                "inspection_focus": to_str(d.get("inspection_focus")),
                "inspection_time": to_str(d.get("inspection_time")),
                "inspection_result": to_str(d.get("inspection_result")),
            }
            cols = list(fields.keys())
            sets = ",".join(["%s=?" % col for col in cols])
            c = conn_db()
            c.execute("INSERT INTO proc_meta (product_code, contract_no, %s) VALUES (?,?,%s) "
                      "ON CONFLICT(product_code, contract_no) DO UPDATE SET %s"
                      % (",".join(cols), ",".join(["?"] * len(cols)), sets),
                      [pc, cn] + [fields[col] for col in cols] + [fields[col] for col in cols])
            c.commit(); c.close()
            send_json(self, {"ok": True})
            return
        self.send_error(404)

    def do_PUT(self):
        u = urlparse(self.path)
        p = u.path
        if p.startswith("/api/raw/"):
            s = require_login(self)
            if not s:
                return
            if not can_write(self, s, p):
                send_json(self, {"error": "无权限"}, 403)
                return
            rid = p.split("/")[-1]
            d = read_body(self)
            errs = validate_raw(d)
            if errs:
                send_json(self, {"error": "校验失败", "details": errs}, 400)
                return
            cd = clean_raw(d)
            cols = [c for c in RAW_COLS if c in cd and cd[c] is not None]
            sets = ",".join(["%s=?" % c for c in cols])
            vals = [cd.get(c) for c in cols] + [rid]
            c = conn_db()
            c.execute("UPDATE raw_orders SET %s, updated_at=datetime('now','localtime') WHERE id=?" % sets, vals)
            c.commit(); c.close()
            send_json(self, {"ok": True})
            return
        self.send_error(404)

    def do_DELETE(self):
        u = urlparse(self.path)
        p = u.path
        if p == "/api/raw/batch":
            s = require_login(self)
            if not s:
                return
            if not can_write(self, s, p):
                send_json(self, {"error": "无权限（删除仅管理员）"}, 403)
                return
            d = read_body(self)
            if not isinstance(d, dict):
                d = {}
            ids = [int(x) for x in d.get("ids", []) if str(x).strip().isdigit()]
            # 跨页全选：按当前搜索/筛选条件整批删除所有匹配行
            if d.get("scope") == "all":
                search = (d.get("search") or "").strip()
                flt = {col: (d.get(col) or "").strip() for col in FILTER_COLS}
                flt = {k: v for k, v in flt.items() if v}
                where, args = build_raw_where(search, flt)
                c = conn_db()
                n = c.execute("DELETE FROM raw_orders " + where, args).rowcount
                c.commit(); c.close()
                send_json(self, {"ok": True, "deleted": n, "scope": "all"})
                return
            if not ids:
                send_json(self, {"error": "未提供有效 ID"}, 400)
                return
            c = conn_db()
            cur = c.execute("DELETE FROM raw_orders WHERE id IN (%s)" % ",".join("?" * len(ids)), ids)
            n = cur.rowcount
            c.commit(); c.close()
            send_json(self, {"ok": True, "deleted": n})
            return
        if p.startswith("/api/raw/") or p.startswith("/api/products/") or \
           p.startswith("/api/shipping/") or p.startswith("/api/forecast/"):
            s = require_login(self)
            if not s:
                return
            if not can_write(self, s, p):
                send_json(self, {"error": "无权限（删除仅管理员）"}, 403)
                return
            rid = p.split("/")[-1]
            if p.startswith("/api/raw/"):
                tname = "raw_orders"
            else:
                tname = {"products": "products", "shipping": "shipping_demands",
                         "forecast": "sales_forecast"}[p.split("/")[2]]
            c = conn_db()
            c.execute("DELETE FROM %s WHERE id=?" % tname, (rid,))
            c.commit(); c.close()
            send_json(self, {"ok": True})
            return
        self.send_error(404)


def main():
    init_db()
    backup_db()
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print("跟单网页原型已启动： http://%s:%d" % (HOST, PORT))
    print("按 Ctrl+C 停止")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        server.shutdown()


if __name__ == "__main__":
    main()
